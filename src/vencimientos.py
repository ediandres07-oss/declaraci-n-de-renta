"""Gestor de vencimientos tributarios para contadores.

Un contador registra sus clientes (NIT + obligaciones que le aplican) y la app
calcula todas las fechas límite del año a partir del calendario oficial DIAN
(config/calendario_<año>.yaml). Ofrece vista calendario, exportación .ics para
Google/Outlook/Apple Calendar e informe PDF de próximos vencimientos.

Gratis con solo iniciar sesión (gancho del canal de contadores): no exige el
pase de temporada.
"""
import hashlib
import hmac
import io
import json
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

import yaml
from flask import (Blueprint, Response, current_app, jsonify, render_template,
                   request, send_file)

from src.auth import db, login_requerido, usuario_actual

BASE = Path(__file__).resolve().parent.parent
CONFIG_DIR = BASE / "config"

venc_bp = Blueprint("vencimientos", __name__)

_CAL_CACHE: Dict[int, dict] = {}


# ---------------------------------------------------------------- modelos
class ClienteContador(db.Model):
    """Cliente de un contador (no confundir con Usuario: el cliente no entra
    a la app; es un registro del portafolio del contador que sí tiene sesión)."""
    __tablename__ = "clientes_contador"
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey("usuarios.id"), index=True)
    nombre = db.Column(db.String(200))
    nit = db.Column(db.String(30))
    tipo = db.Column(db.String(20), default="natural")   # natural|juridica|gran|rst
    correo = db.Column(db.String(200))
    telefono = db.Column(db.String(40))
    notas = db.Column(db.Text)
    obligaciones = db.Column(db.Text)                    # JSON: ["renta_pn", ...]
    creado = db.Column(db.DateTime, default=datetime.utcnow)

    def lista_obligaciones(self) -> List[str]:
        try:
            datos = json.loads(self.obligaciones or "[]")
            return [str(x) for x in datos] if isinstance(datos, list) else []
        except Exception:
            return []

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "nombre": self.nombre,
            "nit": self.nit,
            "tipo": self.tipo,
            "correo": self.correo,
            "telefono": self.telefono,
            "notas": self.notas,
            "obligaciones": self.lista_obligaciones(),
        }


class PerfilContador(db.Model):
    """Marca del contador para el informe PDF: nombre del estudio y logo."""
    __tablename__ = "perfiles_contador"
    usuario_id = db.Column(db.Integer, db.ForeignKey("usuarios.id"),
                           primary_key=True)
    estudio = db.Column(db.String(200))
    logo = db.Column(db.LargeBinary)             # PNG/JPG, máx ~500 KB
    logo_mime = db.Column(db.String(40))
    actualizado = db.Column(db.DateTime, default=datetime.utcnow)


class VencimientoAviso(db.Model):
    """Recordatorio ya enviado al contador (evita duplicados). La clave es el
    evento completo: "<cliente_id>|<clave evento>|<días de anticipo>". También
    guarda el candado diario del lote (usuario_id=0, clave "lote|<fecha>") con
    el que los workers de gunicorn se reparten quién envía hoy — por eso
    usuario_id NO es llave foránea."""
    __tablename__ = "vencimientos_avisos"
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, index=True)
    clave = db.Column(db.String(200))
    enviado = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (db.UniqueConstraint("usuario_id", "clave"),)


class VencimientoEstado(db.Model):
    """Seguimiento por vencimiento: el contador marca presentada/pagada.
    La clave identifica el evento: "<obligacion>|<etiqueta>|<fecha ISO>"."""
    __tablename__ = "vencimientos_estado"
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey("clientes_contador.id"), index=True)
    clave = db.Column(db.String(160))
    estado = db.Column(db.String(20))                    # presentada | pagada
    actualizado = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (db.UniqueConstraint("cliente_id", "clave"),)


# ---------------------------------------------------------------- calendario
def cargar_calendario(ano: int = 2026) -> dict:
    """Lee config/calendario_<año>.yaml (con caché en memoria)."""
    if ano not in _CAL_CACHE:
        ruta = CONFIG_DIR / f"calendario_{ano}.yaml"
        with open(ruta, encoding="utf-8") as fh:
            _CAL_CACHE[ano] = yaml.safe_load(fh)
    return _CAL_CACHE[ano]


def _clave_aplica(clave: str, digito: int, ancho: int) -> bool:
    """¿La clave de fechas ("3", "1-2", "01-05", "9-0", "todos") cubre el dígito?

    El dígito 0 va al final del ciclo DIAN, así que 0→10 (y "00"→100) para que
    rangos como "9-0" o "96-00" queden contiguos.
    """
    if clave == "todos":
        return True
    tope = 10 ** ancho
    valor = digito if digito != 0 else tope
    if "-" in clave:
        ini, fin = clave.split("-", 1)
        a, b = int(ini), int(fin)
        if a == 0:
            a = tope
        if b == 0:
            b = tope
        return a <= valor <= b
    exacto = int(clave)
    if exacto == 0:
        exacto = tope
    return valor == exacto


def _digito_de(nit: str, ancho: int) -> Optional[int]:
    digitos = "".join(c for c in str(nit or "") if c.isdigit())
    if len(digitos) < ancho:
        return None
    return int(digitos[-ancho:])


def vencimientos_de(nit: str, obligaciones: List[str], ano: int = 2026) -> List[dict]:
    """Todos los vencimientos del año para un NIT según sus obligaciones."""
    cal = cargar_calendario(ano)
    catalogo = cal.get("obligaciones", {})
    eventos = []
    for clave_ob in obligaciones:
        ob = catalogo.get(clave_ob)
        if not ob:
            continue
        ancho = int(ob.get("digitos", 1)) or 1
        digito = _digito_de(nit, ancho)
        if digito is None:
            continue
        for periodo in ob.get("periodos", []):
            for clave_fecha, fecha in (periodo.get("fechas") or {}).items():
                if _clave_aplica(str(clave_fecha), digito, ancho):
                    eventos.append({
                        "obligacion": clave_ob,
                        "nombre": ob.get("nombre", clave_ob),
                        "etiqueta": periodo.get("etiqueta", ""),
                        "fecha": fecha if isinstance(fecha, date) else
                                 datetime.strptime(str(fecha), "%Y-%m-%d").date(),
                    })
                    break
    eventos.sort(key=lambda e: e["fecha"])
    return eventos


def catalogo_obligaciones(ano: int = 2026) -> List[dict]:
    cal = cargar_calendario(ano)
    return [{"clave": k, "nombre": v.get("nombre", k)}
            for k, v in cal.get("obligaciones", {}).items()]


# Obligaciones sugeridas al elegir el tipo de cliente (el contador puede ajustar)
SUGERENCIAS = {
    "natural": ["renta_pn"],
    "juridica": ["renta_pj", "iva_bimestral", "retefuente", "exogena"],
    "gran": ["renta_gc", "iva_bimestral", "retefuente", "exogena_gc"],
    "rst": ["rst", "rst_anticipos", "exogena"],
}


# ---------------------------------------------------------------- helpers web
def _cliente_del_usuario(cliente_id: int) -> Optional[ClienteContador]:
    u = usuario_actual()
    if u is None:
        return None
    return ClienteContador.query.filter_by(id=cliente_id, usuario_id=u.id).first()


def _estados_por_cliente(cliente_ids: List[int]) -> Dict[int, Dict[str, str]]:
    if not cliente_ids:
        return {}
    filas = VencimientoEstado.query.filter(
        VencimientoEstado.cliente_id.in_(cliente_ids)).all()
    resultado: Dict[int, Dict[str, str]] = {}
    for f in filas:
        resultado.setdefault(f.cliente_id, {})[f.clave] = f.estado
    return resultado


def _clave_evento(ev: dict) -> str:
    return f"{ev['obligacion']}|{ev['etiqueta']}|{ev['fecha'].isoformat()}"


def _agenda_usuario(u) -> List[dict]:
    """Todos los eventos de todos los clientes del contador, con estado."""
    clientes = ClienteContador.query.filter_by(usuario_id=u.id).all()
    estados = _estados_por_cliente([c.id for c in clientes])
    agenda = []
    for c in clientes:
        for ev in vencimientos_de(c.nit, c.lista_obligaciones()):
            clave = _clave_evento(ev)
            agenda.append({
                "cliente_id": c.id,
                "cliente": c.nombre,
                "nit": c.nit,
                "obligacion": ev["obligacion"],
                "nombre": ev["nombre"],
                "etiqueta": ev["etiqueta"],
                "fecha": ev["fecha"].isoformat(),
                "clave": clave,
                "estado": estados.get(c.id, {}).get(clave, "pendiente"),
            })
    agenda.sort(key=lambda e: (e["fecha"], e["cliente"] or ""))
    return agenda


# ---------------------------------------------------------------- rutas página
@venc_bp.route("/vencimientos")
@login_requerido
def pagina_vencimientos():
    resp = Response(render_template("vencimientos.html",
                                    ano=cargar_calendario().get("ano", 2026)))
    # Safari sirve versiones viejas de la página tras un despliegue; que no cachee
    resp.headers["Cache-Control"] = "no-store"
    return resp


# ---------------------------------------------------------------- rutas API
@venc_bp.route("/api/vencimientos/catalogo")
@login_requerido
def api_catalogo():
    return jsonify({"ano": cargar_calendario().get("ano", 2026),
                    "obligaciones": catalogo_obligaciones(),
                    "sugerencias": SUGERENCIAS})


@venc_bp.route("/api/vencimientos/clientes", methods=["GET", "POST"])
@login_requerido
def api_clientes():
    u = usuario_actual()
    if request.method == "POST":
        datos = request.get_json(silent=True) or {}
        nombre = (datos.get("nombre") or "").strip()
        nit = "".join(c for c in str(datos.get("nit") or "") if c.isdigit())
        if not nombre or len(nit) < 2:
            return jsonify({"error": "Nombre y NIT/cédula (mínimo 2 dígitos) son "
                            "obligatorios."}), 400
        obligaciones = datos.get("obligaciones")
        if not isinstance(obligaciones, list) or not obligaciones:
            obligaciones = SUGERENCIAS.get(datos.get("tipo") or "natural",
                                           ["renta_pn"])
        validas = {o["clave"] for o in catalogo_obligaciones()}
        obligaciones = [o for o in obligaciones if o in validas]
        cliente = ClienteContador(
            usuario_id=u.id, nombre=nombre[:200], nit=nit[:30],
            tipo=(datos.get("tipo") or "natural")[:20],
            correo=(datos.get("correo") or "")[:200] or None,
            telefono=(datos.get("telefono") or "")[:40] or None,
            notas=(datos.get("notas") or "")[:2000] or None,
            obligaciones=json.dumps(obligaciones))
        db.session.add(cliente)
        db.session.commit()
        return jsonify({"ok": True, "cliente": cliente.to_dict()})
    clientes = (ClienteContador.query.filter_by(usuario_id=u.id)
                .order_by(ClienteContador.nombre).all())
    hoy = date.today()
    respuesta = []
    for c in clientes:
        d = c.to_dict()
        proximos = [e for e in vencimientos_de(c.nit, c.lista_obligaciones())
                    if e["fecha"] >= hoy]
        d["proximo"] = ({"nombre": proximos[0]["nombre"],
                         "etiqueta": proximos[0]["etiqueta"],
                         "fecha": proximos[0]["fecha"].isoformat()}
                        if proximos else None)
        respuesta.append(d)
    return jsonify({"clientes": respuesta})


@venc_bp.route("/api/vencimientos/clientes/<int:cliente_id>",
               methods=["PUT", "DELETE"])
@login_requerido
def api_cliente(cliente_id):
    cliente = _cliente_del_usuario(cliente_id)
    if cliente is None:
        return jsonify({"error": "Cliente no encontrado."}), 404
    if request.method == "DELETE":
        VencimientoEstado.query.filter_by(cliente_id=cliente.id).delete()
        db.session.delete(cliente)
        db.session.commit()
        return jsonify({"ok": True})
    datos = request.get_json(silent=True) or {}
    if "nombre" in datos:
        nombre = (datos.get("nombre") or "").strip()
        if not nombre:
            return jsonify({"error": "El nombre no puede quedar vacío."}), 400
        cliente.nombre = nombre[:200]
    if "nit" in datos:
        nit = "".join(c for c in str(datos.get("nit") or "") if c.isdigit())
        if len(nit) < 2:
            return jsonify({"error": "NIT/cédula inválido."}), 400
        cliente.nit = nit[:30]
    if "tipo" in datos:
        cliente.tipo = (datos.get("tipo") or "natural")[:20]
    for campo, tope in (("correo", 200), ("telefono", 40), ("notas", 2000)):
        if campo in datos:
            setattr(cliente, campo, (datos.get(campo) or "")[:tope] or None)
    if "obligaciones" in datos and isinstance(datos["obligaciones"], list):
        validas = {o["clave"] for o in catalogo_obligaciones()}
        cliente.obligaciones = json.dumps(
            [o for o in datos["obligaciones"] if o in validas])
    db.session.commit()
    return jsonify({"ok": True, "cliente": cliente.to_dict()})


@venc_bp.route("/api/vencimientos/agenda")
@login_requerido
def api_agenda():
    u = usuario_actual()
    return jsonify({"hoy": date.today().isoformat(), "eventos": _agenda_usuario(u)})


@venc_bp.route("/api/vencimientos/estado", methods=["POST"])
@login_requerido
def api_estado():
    datos = request.get_json(silent=True) or {}
    cliente = _cliente_del_usuario(int(datos.get("cliente_id") or 0))
    if cliente is None:
        return jsonify({"error": "Cliente no encontrado."}), 404
    clave = (datos.get("clave") or "")[:160]
    estado = datos.get("estado") or "pendiente"
    if estado not in ("pendiente", "presentada", "pagada"):
        return jsonify({"error": "Estado inválido."}), 400
    fila = VencimientoEstado.query.filter_by(cliente_id=cliente.id,
                                             clave=clave).first()
    if estado == "pendiente":
        if fila:
            db.session.delete(fila)
    elif fila:
        fila.estado = estado
        fila.actualizado = datetime.utcnow()
    else:
        db.session.add(VencimientoEstado(cliente_id=cliente.id, clave=clave,
                                         estado=estado))
    db.session.commit()
    return jsonify({"ok": True, "estado": estado})


# ---------------------------------------------------------------- iCalendar
def _firma_ical(usuario_id: int) -> str:
    secreto = current_app.secret_key or ""
    return hmac.new(str(secreto).encode(), f"ical:{usuario_id}".encode(),
                    hashlib.sha256).hexdigest()[:24]


def _ics_de_eventos(eventos: List[dict], nombre_cal: str) -> str:
    """Arma un VCALENDAR simple (eventos de día completo + alarma 3 días antes)."""
    lineas = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//tributando.co//vencimientos//ES",
        "CALSCALE:GREGORIAN",
        f"X-WR-CALNAME:{nombre_cal}",
    ]
    for ev in eventos:
        fecha = ev["fecha"].replace("-", "")
        uid = hashlib.sha1(f"{ev['cliente_id']}|{ev['clave']}".encode()).hexdigest()
        resumen = f"{ev['cliente']}: {ev['nombre']}"
        if ev["etiqueta"]:
            resumen += f" — {ev['etiqueta']}"
        resumen = resumen.replace(",", "\\,").replace(";", "\\;")
        lineas += [
            "BEGIN:VEVENT",
            f"UID:{uid}@tributando.co",
            f"DTSTART;VALUE=DATE:{fecha}",
            f"SUMMARY:{resumen}",
            f"DESCRIPTION:NIT {ev['nit']} · Vencimiento DIAN. "
            "Gestor de vencimientos de tributando.co",
            "BEGIN:VALARM",
            "ACTION:DISPLAY",
            "DESCRIPTION:Vencimiento tributario en 3 días",
            "TRIGGER:-P3D",
            "END:VALARM",
            "END:VEVENT",
        ]
    lineas.append("END:VCALENDAR")
    return "\r\n".join(lineas) + "\r\n"


@venc_bp.route("/api/vencimientos/ical")
@login_requerido
def api_ical():
    """Descarga .ics (todos los clientes o ?cliente=<id>) + URL de suscripción."""
    u = usuario_actual()
    eventos = _agenda_usuario(u)
    cliente_id = request.args.get("cliente", type=int)
    if cliente_id:
        eventos = [e for e in eventos if e["cliente_id"] == cliente_id]
    if request.args.get("url"):
        base = request.url_root.rstrip("/")
        return jsonify({"url": f"{base}/vencimientos/ical/"
                               f"{u.id}-{_firma_ical(u.id)}.ics"})
    ics = _ics_de_eventos(eventos, "Vencimientos tributarios")
    return send_file(io.BytesIO(ics.encode("utf-8")),
                     mimetype="text/calendar",
                     as_attachment=True,
                     download_name="vencimientos-tributando.ics")


@venc_bp.route("/vencimientos/ical/<int:usuario_id>-<firma>.ics")
def ical_suscripcion(usuario_id, firma):
    """URL de suscripción (sin sesión: la usan Google/Apple Calendar). La firma
    HMAC del id evita que un tercero adivine el calendario de otro contador."""
    if not hmac.compare_digest(firma, _firma_ical(usuario_id)):
        return Response("No autorizado", status=403)
    from src.auth import Usuario
    u = db.session.get(Usuario, usuario_id)
    if u is None:
        return Response("No existe", status=404)
    ics = _ics_de_eventos(_agenda_usuario(u), "Vencimientos tributarios")
    return Response(ics, mimetype="text/calendar")


# ---------------------------------------------------------------- informe PDF
@venc_bp.route("/api/vencimientos/informe.pdf")
@login_requerido
def api_informe():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import (HRFlowable, Image, Paragraph,
                                    SimpleDocTemplate, Spacer, Table, TableStyle)

    NAVY = colors.HexColor("#1e2432")
    DORADO = colors.HexColor("#cdab7e")
    GRIS = colors.HexColor("#5a6b7f")

    u = usuario_actual()
    dias = min(max(request.args.get("dias", default=30, type=int), 1), 365)
    hoy = date.today()
    eventos = [e for e in _agenda_usuario(u)
               if 0 <= (date.fromisoformat(e["fecha"]) - hoy).days <= dias
               and e["estado"] == "pendiente"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=18 * mm,
                            bottomMargin=18 * mm, leftMargin=18 * mm,
                            rightMargin=18 * mm,
                            title="Informe de vencimientos")
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle("t", parent=estilos["Title"], textColor=NAVY,
                            fontSize=18, spaceAfter=2)
    sub = ParagraphStyle("s", parent=estilos["Normal"], textColor=GRIS,
                         fontSize=10, spaceAfter=8)
    celda = ParagraphStyle("c", parent=estilos["Normal"], fontSize=9,
                           textColor=NAVY)

    # marca del contador (logo + nombre del estudio), si la configuró
    perfil = db.session.get(PerfilContador, u.id)
    firma = f" · {perfil.estudio}" if (perfil and perfil.estudio) else ""
    piezas = []
    if perfil and perfil.logo:
        try:
            img = ImageReader(io.BytesIO(perfil.logo))
            ancho_px, alto_px = img.getSize()
            alto = 16 * mm
            piezas += [Image(io.BytesIO(perfil.logo),
                             width=alto * ancho_px / max(alto_px, 1),
                             height=alto, hAlign="LEFT"),
                       Spacer(1, 3 * mm)]
        except Exception:
            pass                                  # logo corrupto: seguimos sin él
    piezas += [Paragraph("Informe de vencimientos tributarios", titulo),
               Paragraph(f"Próximos {dias} días · generado el "
                         f"{hoy.strftime('%d/%m/%Y')}{firma} · tributando.co", sub),
               HRFlowable(width="100%", thickness=2, color=DORADO),
               Spacer(1, 6 * mm)]

    if not eventos:
        piezas.append(Paragraph("Sin vencimientos pendientes en el periodo. 🎉",
                                estilos["Normal"]))
    else:
        filas = [[Paragraph(f"<b>{c}</b>", celda) for c in
                  ("Fecha", "Cliente", "NIT", "Obligación", "Días")]]
        for e in eventos:
            f = date.fromisoformat(e["fecha"])
            restantes = (f - hoy).days
            oblig = e["nombre"] + (f" — {e['etiqueta']}" if e["etiqueta"] else "")
            filas.append([Paragraph(f.strftime("%d/%m/%Y"), celda),
                          Paragraph(e["cliente"] or "", celda),
                          Paragraph(e["nit"] or "", celda),
                          Paragraph(oblig, celda),
                          Paragraph(str(restantes), celda)])
        tabla = Table(filas, colWidths=[24 * mm, 48 * mm, 26 * mm, 66 * mm,
                                        14 * mm], repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f4f0e9")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d8d3c8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        # el encabezado usa Paragraph con estilo navy; píntalo blanco a mano
        for p in filas[0]:
            p.style = ParagraphStyle("h", parent=celda, textColor=colors.white)
        piezas.append(tabla)

    doc.build(piezas)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"informe-vencimientos-{hoy.isoformat()}.pdf")


# ---------------------------------------------------------------- importar Excel
COLUMNAS_PLANTILLA = ["nombre", "nit", "tipo", "correo", "telefono", "obligaciones"]


@venc_bp.route("/api/vencimientos/plantilla.xlsx")
@login_requerido
def api_plantilla_importar():
    """Plantilla de importación con ejemplos y la lista de códigos válidos."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(COLUMNAS_PLANTILLA)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1E2432")
    ws.append(["Panadería La Espiga SAS", "901234567", "juridica",
               "espiga@ejemplo.com", "3001234567", "renta_pj,iva_bimestral,retefuente"])
    ws.append(["Ana María Pérez", "1030601234", "natural", "", "", ""])
    for col, ancho in zip("ABCDEF", (32, 16, 12, 26, 14, 40)):
        ws.column_dimensions[col].width = ancho
    guia = wb.create_sheet("Códigos")
    guia.append(["Código de obligación", "Nombre"])
    guia["A1"].font = guia["B1"].font = Font(bold=True)
    for o in catalogo_obligaciones():
        guia.append([o["clave"], o["nombre"]])
    guia.append([])
    guia.append(["tipo:", "natural | juridica | gran | rst"])
    guia.append(["obligaciones:", "códigos separados por coma; si se deja vacío "
                 "se asignan las sugeridas según el tipo"])
    guia.column_dimensions["A"].width = 24
    guia.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="plantilla-clientes-tributando.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


def _filas_de_archivo(archivo) -> List[dict]:
    """Lee .xlsx o .csv y devuelve dicts con las columnas de la plantilla."""
    nombre = (archivo.filename or "").lower()
    if nombre.endswith(".csv"):
        import csv
        texto = archivo.read().decode("utf-8-sig", errors="replace")
        return list(csv.DictReader(texto.splitlines()))
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True,
                                read_only=True)
    ws = wb.worksheets[0]
    filas = ws.iter_rows(values_only=True)
    encabezado = [str(c or "").strip().lower() for c in next(filas, [])]
    datos = []
    for fila in filas:
        d = {encabezado[i]: fila[i] for i in range(min(len(encabezado), len(fila)))}
        datos.append(d)
    wb.close()
    return datos


@venc_bp.route("/api/vencimientos/importar", methods=["POST"])
@login_requerido
def api_importar():
    u = usuario_actual()
    archivo = request.files.get("archivo")
    if archivo is None:
        return jsonify({"error": "Adjunta un archivo .xlsx o .csv."}), 400
    try:
        filas = _filas_de_archivo(archivo)
    except Exception:
        return jsonify({"error": "No pude leer el archivo. Usa la plantilla "
                        ".xlsx o un .csv con encabezados."}), 400
    validas = {o["clave"] for o in catalogo_obligaciones()}
    existentes = {(c.nit or "") for c in
                  ClienteContador.query.filter_by(usuario_id=u.id).all()}
    creados, errores = 0, []
    for i, fila in enumerate(filas, start=2):     # 2 = primera fila de datos
        nombre = str(fila.get("nombre") or "").strip()
        nit = "".join(c for c in str(fila.get("nit") or "") if c.isdigit())
        if not nombre and not nit:
            continue                               # fila vacía
        if not nombre or len(nit) < 2:
            errores.append(f"Fila {i}: falta nombre o NIT válido.")
            continue
        if nit in existentes:
            errores.append(f"Fila {i}: el NIT {nit} ya existe, se omitió.")
            continue
        tipo = str(fila.get("tipo") or "natural").strip().lower()
        if tipo not in SUGERENCIAS:
            tipo = "natural"
        crudas = [x.strip() for x in
                  str(fila.get("obligaciones") or "").replace(";", ",").split(",")
                  if x.strip()]
        obligaciones = [o for o in crudas if o in validas] or SUGERENCIAS[tipo]
        db.session.add(ClienteContador(
            usuario_id=u.id, nombre=nombre[:200], nit=nit[:30], tipo=tipo,
            correo=str(fila.get("correo") or "")[:200] or None,
            telefono=str(fila.get("telefono") or "")[:40] or None,
            obligaciones=json.dumps(obligaciones)))
        existentes.add(nit)
        creados += 1
    db.session.commit()
    return jsonify({"ok": True, "creados": creados, "errores": errores[:20]})


# ---------------------------------------------------------------- leer RUT
# Responsabilidades del RUT (casilla 53) → obligaciones del calendario.
RUT_CODIGOS = {
    "03": ["patrimonio"],
    "05": [],                       # renta ordinaria: PJ o PN según el tipo
    "07": ["retefuente"],           # agente retenedor renta
    "08": ["retefuente"],           # retención timbre
    "09": ["retefuente"],           # agente retenedor IVA
    "13": ["renta_gc", "exogena_gc"],
    "14": ["exogena"],
    "15": ["retefuente"],           # autorretenedor
    "18": ["precios_transferencia"],
    "47": ["rst", "rst_anticipos"],
    "48": ["iva_bimestral"],
}


def analizar_rut(texto: str) -> dict:
    """Extrae NIT, nombre, tipo y obligaciones sugeridas del texto de un RUT."""
    import re

    plano = " ".join(texto.split())

    def solo_digitos(t):
        return "".join(c for c in t if c.isdigit())

    # En el RUT oficial pypdf saca primero TODAS las etiquetas del formulario y
    # después los VALORES (con los dígitos sueltos: "7 2 2 2 1 6 3"). Por eso el
    # NIT se ancla a los valores, no a la etiqueta "(NIT)".
    nit = ""
    # 1) persona natural: la cédula va tras "Cédula de Ciudadanía" + código 13
    m = re.search(r"C[eé]dula de Ciudadan[ií]a\s+1\s+3\s+((?:\d\s+){4,10}\d)",
                  plano)
    if m:
        nit = solo_digitos(m.group(1))
    # 2) tras el número de formulario (contiguo, 10-14 dígitos) viene el NIT
    if not nit:
        m = re.search(r"\b\d{10,14}\b\s+((?:\d\s+){4,10}\d)", plano)
        if m:
            nit = solo_digitos(m.group(1))
            if len(nit) == 10 and nit[0] in "89":
                nit = nit[:-1]                  # venía con el DV pegado
    # 3) respaldo: PDFs que sí traen dígitos junto a la etiqueta "(NIT)"
    if not nit:
        sin_casillas = re.sub(r"\b\d{1,2}\.\s", " ", plano)
        m = re.search(r"Identificaci[oó]n Tributaria \(NIT\)[^0-9]{0,40}"
                      r"(\d[\d .]{4,20}\d)", sin_casillas)
        if m:
            nit = solo_digitos(m.group(1))
            m_dv = re.search(r"DV[^0-9]{0,10}(\d)\b", plano)
            if m_dv and nit.endswith(m_dv.group(1)) and len(nit) >= 9:
                nit = nit[:-1]

    juridica = bool(re.search(r"Persona jur[ií]dica", plano, re.I))
    natural = bool(re.search(r"Persona natural", plano, re.I))

    nombre = ""
    # En el RUT oficial el nombre en MAYÚSCULAS va justo antes de la calidad
    # con la que firma ("... ROJAS ACERO GUILLERMO ADOLFO CONTRIBUYENTE")
    m = re.search(r"([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ ]{6,70})\s+"
                  r"(?:CONTRIBUYENTE|REPRESENTANTE LEGAL|APODERADO)", plano)
    if m:
        palabras = [p for p in m.group(1).split()
                    if p not in ("AM", "PM", "X", "SI", "NO") and len(p) > 1]
        nombre = " ".join(palabras)
    if not nombre:
        # RUT sin firma (actuación de oficio): el nombre en mayúsculas va justo
        # antes del país de la dirección ("... EDISON ANDRES COLOMBIA 1 6 9")
        m = re.search(r"([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ ]{6,70})\s+COLOMBIA\s+1\s+6\s+9",
                      plano)
        if m:
            nombre = " ".join(m.group(1).split())
    if not nombre:
        m = re.search(r"raz[oó]n social:?\s*(.{3,90}?)\s*(?:3[67]\.|Nombre "
                      r"comercial|Sigla|$)", plano, re.I)
        if m and m.group(1).strip("- ").strip():
            nombre = m.group(1).strip("- ").strip()
    # Antes que llenar el campo con etiquetas del formulario ("32. Segundo
    # apellido"...), mejor dejarlo vacío.
    if nombre and (re.match(r"^\d", nombre) or
                   re.search(r"apellido|nombre|raz[oó]n social|sigla", nombre, re.I)):
        nombre = ""

    codigos = set()
    # forma "05- Impto. renta..." o "48 Impuesto sobre las ventas"
    for m in re.finditer(r"\b(\d{2})\s*-\s*[A-ZÁÉÍÓÚa-z]", plano):
        if m.group(1) in RUT_CODIGOS:
            codigos.add(m.group(1))
    # respaldo por frases (por si la casilla 53 no trae los números legibles)
    frases = {
        "impuesto al patrimonio": "03",
        "r[eé]gimen ordinario": "05",
        "retenci[oó]n en la fuente a t[ií]tulo de renta": "07",
        "gran contribuyente": "13",
        "informante de ex[oó]gena": "14",
        "autorretenedor": "15",
        "precios de transferencia": "18",
        "r[eé]gimen simple": "47",
        "impuesto sobre las ventas": "48",
    }
    for patron, codigo in frases.items():
        if re.search(patron, plano, re.I):
            codigos.add(codigo)

    # El formulario del RUT trae impresas AMBAS frases ("Persona natural" y
    # "Persona jurídica") como etiquetas, así que el NIT decide primero: los de
    # sociedades tienen 9 dígitos y empiezan por 8 o 9.
    if nit:
        tipo = "juridica" if (len(nit) == 9 and nit[0] in "89") else "natural"
    else:
        tipo = "juridica" if (juridica and not natural) else "natural"
    if "13" in codigos:
        tipo = "gran"
    elif "47" in codigos:
        tipo = "rst"
    obligaciones = set()
    for c in codigos:
        obligaciones.update(RUT_CODIGOS[c])
    if "05" in codigos:
        obligaciones.add({"juridica": "renta_pj", "gran": "renta_gc",
                          "rst": "rst"}.get(tipo, "renta_pn"))
    if not obligaciones:
        obligaciones.update(SUGERENCIAS.get(tipo, []))
    validas = {o["clave"] for o in catalogo_obligaciones()}
    return {"nit": nit, "nombre": nombre, "tipo": tipo,
            "codigos": sorted(codigos),
            "obligaciones": sorted(o for o in obligaciones if o in validas)}


@venc_bp.route("/api/vencimientos/rut", methods=["POST"])
@login_requerido
def api_leer_rut():
    archivo = request.files.get("archivo")
    if archivo is None:
        return jsonify({"error": "Adjunta el RUT en PDF."}), 400
    try:
        import pypdf
        lector = pypdf.PdfReader(io.BytesIO(archivo.read()))
        texto = "\n".join((p.extract_text() or "") for p in lector.pages[:3])
    except Exception:
        return jsonify({"error": "No pude leer ese PDF."}), 400
    datos = analizar_rut(texto)
    if not datos["nit"] and not datos["codigos"]:
        return jsonify({"error": "Ese PDF no parece un RUT con texto legible "
                        "(¿es un escaneo?). Ingresa los datos a mano."}), 422
    return jsonify({"ok": True, **datos})


# ---------------------------------------------------------------- perfil (logo)
@venc_bp.route("/api/vencimientos/perfil", methods=["GET", "POST"])
@login_requerido
def api_perfil():
    u = usuario_actual()
    perfil = db.session.get(PerfilContador, u.id)
    if request.method == "GET":
        return jsonify({"estudio": perfil.estudio if perfil else "",
                        "tiene_logo": bool(perfil and perfil.logo)})
    if perfil is None:
        perfil = PerfilContador(usuario_id=u.id)
        db.session.add(perfil)
    if "estudio" in request.form:
        perfil.estudio = (request.form.get("estudio") or "")[:200] or None
    logo = request.files.get("logo")
    if logo is not None:
        datos = logo.read()
        if len(datos) > 500 * 1024:
            return jsonify({"error": "El logo no puede pesar más de 500 KB."}), 400
        mime = (logo.mimetype or "").lower()
        if mime not in ("image/png", "image/jpeg"):
            return jsonify({"error": "El logo debe ser PNG o JPG."}), 400
        try:                                     # que no rompa el PDF después
            from PIL import Image as PILImage
            PILImage.open(io.BytesIO(datos)).verify()
        except Exception:
            return jsonify({"error": "Ese archivo de imagen no se pudo leer; "
                            "exporta el logo de nuevo como PNG o JPG."}), 400
        perfil.logo, perfil.logo_mime = datos, mime
    perfil.actualizado = datetime.utcnow()
    db.session.commit()
    return jsonify({"ok": True, "estudio": perfil.estudio or "",
                    "tiene_logo": bool(perfil.logo)})


@venc_bp.route("/api/vencimientos/perfil/logo")
@login_requerido
def api_perfil_logo():
    u = usuario_actual()
    perfil = db.session.get(PerfilContador, u.id)
    if not (perfil and perfil.logo):
        return Response("Sin logo", status=404)
    return Response(perfil.logo, mimetype=perfil.logo_mime or "image/png")


# ---------------------------------------------------------------- recordatorios
DIAS_AVISO = (7, 3)


def recordatorios_vencimientos(hoy: Optional[date] = None) -> List[dict]:
    """Pendientes de aviso por contador: eventos a 7 o 3 días sin correo previo.

    Devuelve [{usuario, eventos: [(evento, dias_restantes)]}] listo para enviar.
    Debe llamarse dentro de un app context.
    """
    from src.auth import Usuario

    hoy = hoy or date.today()
    pendientes = []
    con_clientes = (db.session.query(ClienteContador.usuario_id)
                    .distinct().all())
    for (uid,) in con_clientes:
        u = db.session.get(Usuario, uid)
        if u is None or not u.email or not u.acepta_recordatorios:
            continue
        avisados = {a.clave for a in
                    VencimientoAviso.query.filter_by(usuario_id=uid).all()}
        eventos = []
        for ev in _agenda_usuario(u):
            if ev["estado"] != "pendiente":
                continue
            restantes = (date.fromisoformat(ev["fecha"]) - hoy).days
            if restantes not in DIAS_AVISO:
                continue
            clave_aviso = f"{ev['cliente_id']}|{ev['clave']}|{restantes}"
            if clave_aviso in avisados:
                continue
            eventos.append({**ev, "dias": restantes, "clave_aviso": clave_aviso})
        if eventos:
            pendientes.append({"usuario": u, "eventos": eventos})
    return pendientes


def plantilla_aviso_vencimientos(nombre: str, eventos: List[dict]) -> tuple:
    """(asunto, html) del correo de aviso al contador."""
    urgentes = [e for e in eventos if e["dias"] <= 3]
    asunto = (f"⚠️ {len(eventos)} vencimiento(s) de tus clientes esta semana"
              if urgentes else
              f"📅 {len(eventos)} vencimiento(s) de tus clientes en 7 días")
    filas = "".join(
        f"<tr><td style='padding:8px 10px;border-bottom:1px solid #e3ddd2'>"
        f"<b>{e['fecha']}</b></td>"
        f"<td style='padding:8px 10px;border-bottom:1px solid #e3ddd2'>{e['cliente']}"
        f"<br><span style='color:#5a6b7f;font-size:12px'>NIT {e['nit']}</span></td>"
        f"<td style='padding:8px 10px;border-bottom:1px solid #e3ddd2'>{e['nombre']}"
        + (f" — {e['etiqueta']}" if e["etiqueta"] else "") +
        f"<br><span style='color:{'#c0392b' if e['dias'] <= 3 else '#b7791f'};"
        f"font-size:12px;font-weight:bold'>en {e['dias']} días</span></td></tr>"
        for e in eventos)
    html = f"""
    <div style="font-family:-apple-system,'Segoe UI',Roboto,sans-serif;max-width:640px;margin:0 auto">
      <div style="background:#1e2432;color:#cdab7e;padding:18px 22px;border-radius:12px 12px 0 0">
        <b style="font-size:17px">Tributando.co</b> · Gestor de vencimientos</div>
      <div style="border:1px solid #e3ddd2;border-top:0;padding:22px;border-radius:0 0 12px 12px">
        <p style="margin:0 0 14px">Hola {nombre or 'contador(a)'}, estos vencimientos de tus
        clientes están cerca:</p>
        <table style="border-collapse:collapse;width:100%;font-size:14px">{filas}</table>
        <p style="margin:18px 0 0"><a href="https://tributando.co/vencimientos"
          style="background:#cdab7e;color:#1e2432;padding:11px 18px;border-radius:10px;
          text-decoration:none;font-weight:bold">Abrir mi calendario →</a></p>
        <p style="color:#8a919c;font-size:12px;margin-top:18px">Recibes este aviso porque
        usas el gestor de vencimientos. Desactívalo en "Mi cuenta".</p>
      </div>
    </div>"""
    return asunto, html


def enviar_avisos_vencimientos(hoy: Optional[date] = None, seco: bool = False) -> int:
    """Envía los avisos pendientes (1 correo por contador). Devuelve enviados.
    Con seco=True solo lista sin enviar ni marcar."""
    from src import correo

    cfg = correo.cargar_config_email()
    enviados = 0
    for p in recordatorios_vencimientos(hoy):
        u, eventos = p["usuario"], p["eventos"]
        if seco:
            print(f"  [seco] {u.email}: {len(eventos)} evento(s)")
            continue
        asunto, html = plantilla_aviso_vencimientos(u.nombre or "", eventos)
        try:
            correo.enviar_email(u.email, asunto, html, cfg)
        except Exception as exc:              # no frenar el lote por un correo
            print(f"  ✗ {u.email}: {exc}")
            continue
        for e in eventos:
            db.session.add(VencimientoAviso(usuario_id=u.id,
                                            clave=e["clave_aviso"]))
        db.session.commit()
        enviados += 1
    return enviados


def correr_avisos_diarios(hoy: Optional[date] = None) -> int:
    """Corrida diaria automática (la llama el hilo de webapp). El candado en BD
    garantiza que un solo worker envíe el lote del día. Devuelve enviados."""
    from src import correo

    hoy = hoy or date.today()
    if not correo.cargar_config_email().get("habilitado"):
        return 0
    try:
        db.session.add(VencimientoAviso(usuario_id=0,
                                        clave=f"lote|{hoy.isoformat()}"))
        db.session.commit()
    except Exception:                     # otro worker ya tomó el lote de hoy
        db.session.rollback()
        return 0
    return enviar_avisos_vencimientos(hoy)
