"""Parser del reporte de Información Exógena de la DIAN.

Lee el .xlsx de "Consulta de información reportada por terceros" y produce
un ResultadoExogena con:
 - metadatos del consultante,
 - partidas clasificadas por renglón del Formulario 210 (columna
   "Uso declaración Sugerida", texto libre y no estandarizado),
 - los 5 "Topes" resumen calculados por la DIAN,
 - advertencias de todo lo que requiera revisión humana.

El parser valida por NOMBRE de columna, no por posición, y tolera
variaciones de formato entre años (más/menos filas de metadatos,
columnas en otro orden, montos como texto).
"""
import datetime as _dt
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

from .modelos import PartidaExogena, ResultadoExogena


class ExogenaError(Exception):
    """Error irrecuperable leyendo el archivo de exógena."""


# --- normalización de texto -------------------------------------------------

def _norm(texto: str) -> str:
    """minúsculas, sin tildes, espacios colapsados — para comparar encabezados."""
    if texto is None:
        return ""
    s = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


# Regex tolerante: "R132", "R 30", "r58", "Renglón 29", "Renglon29", "R29:"
RE_RENGLON = re.compile(r"(?:\bR|\brengl[oó]n)\s*\.?\s*(\d{2,3})\b", re.IGNORECASE)
RE_TOPE = re.compile(r"\btope\s*(\d)\b", re.IGNORECASE)
RE_FECHA = re.compile(r"\d{4}-\d{1,2}-\d{1,2}|\d{1,2}/\d{1,2}/\d{4}|\d{1,2}-\d{1,2}-\d{4}")


def _parse_fechas(texto: str) -> List["_dt.date"]:
    """Extrae fechas de un texto (formatos yyyy-mm-dd, dd/mm/yyyy, dd-mm-yyyy)."""
    out = []
    for s in RE_FECHA.findall(texto or ""):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                out.append(_dt.datetime.strptime(s, fmt).date())
                break
            except ValueError:
                continue
    return out


def _nota_venta_activo(pt: "PartidaExogena") -> str:
    """Nota Art. 300 para una venta de activo fijo. Si la exógena trae las fechas
    de adquisición y venta (en detalle o info adicional), calcula el tiempo de
    posesión y dice si es ganancia ocasional (≥2 años) o renta no laboral (<2)."""
    fechas = sorted(set(_parse_fechas(f"{pt.info_adicional or ''} {pt.detalle or ''}")))
    if len(fechas) >= 2:
        adq, ven = fechas[0], fechas[-1]
        anios = (ven - adq).days / 365.25
        if (ven - adq).days >= 730:
            return (f"Venta de ACTIVO FIJO: adquirido {adq} y vendido {ven} "
                    f"({anios:.1f} años de posesión, ≥2 años) → GANANCIA OCASIONAL "
                    "(R112), correcto (Art. 300).")
        return (f"Venta de ACTIVO FIJO: adquirido {adq} y vendido {ven} "
                f"({anios:.1f} años de posesión, <2 años) → NO es ganancia ocasional: "
                "va a RENTA NO LABORAL. Reclasifique de R112 a R74 (Art. 300).")
    return ("Venta de ACTIVO FIJO. Art. 300: es ganancia ocasional solo si lo poseyó "
            "2 años o más; si <2 años va a rentas no laborales (R74). Verifique la "
            "fecha de adquisición.")

# Beneficiario económico / titularidad (columna "Información Adicional")
RE_PARTICIPACION = re.compile(r"porcentaje\s+de\s+participaci[oó]n:\s*([\d.,]+)", re.IGNORECASE)
RE_PROPIETARIOS = re.compile(r"n[uú]mero\s+(?:de\s+)?propietarios:\s*(\d+)", re.IGNORECASE)
RE_COTITULAR = re.compile(r"titular\s+secundario|cotitular|beneficiario", re.IGNORECASE)

# Encabezados esperados (normalizados) → nombre lógico
_COLUMNAS = {
    "nit": None,  # aparece dos veces; se resuelve por orden
    "nombre / razon social": "informante_nombre",
    "nombre/razon social reportada por el tercero": "informado_nombre",
    "detalle": "detalle",
    "valor": "valor",
    "uso declaracion sugerida": "uso",
    "informacion adicional": "info_adicional",
}


def _parse_valor(v) -> Optional[float]:
    """Convierte el valor de la celda a número; tolera texto con $ , . espacios."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if not s:
        return None
    # formato colombiano 1.234.567,89 o anglosajón 1,234,567.89
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # una coma: decimal si hay <=2 dígitos después, si no separador de miles
        entero, _, dec = s.rpartition(",")
        s = entero.replace(".", "") + ("." + dec if len(dec) <= 2 else dec)
    else:
        # puntos como separador de miles (1.234.567)
        partes = s.split(".")
        if len(partes) > 2 or (len(partes) == 2 and len(partes[1]) == 3):
            s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def _encontrar_encabezado(ws) -> Tuple[int, Dict[str, int]]:
    """Busca la fila de encabezado de la tabla y devuelve (fila, {nombre_lógico: col})."""
    for fila in range(1, min(ws.max_row, 60) + 1):
        valores = [ (_norm(ws.cell(row=fila, column=c).value), c)
                    for c in range(1, ws.max_column + 1)
                    if ws.cell(row=fila, column=c).value is not None ]
        textos = {t for t, _ in valores}
        if "detalle" in textos and "valor" in textos:
            cols: Dict[str, int] = {}
            nits = sorted(c for t, c in valores if t == "nit")
            if len(nits) >= 2:
                cols["informante_nit"], cols["informado_nit"] = nits[0], nits[1]
            elif len(nits) == 1:
                cols["informante_nit"] = nits[0]
            for texto, col in valores:
                for esperado, logico in _COLUMNAS.items():
                    if logico and texto.startswith(esperado):
                        cols.setdefault(logico, col)
            # fallback por posición relativa para los nombres si faltaron
            if "detalle" in cols and "valor" in cols:
                return fila, cols
    raise ExogenaError(
        "No se encontró la fila de encabezado de la tabla (se buscó una fila "
        "con columnas 'Detalle' y 'Valor'). ¿Es este un reporte de exógena de la DIAN?"
    )


def _extraer_metadatos(ws, fila_encabezado: int, resultado: ResultadoExogena) -> None:
    """Lee los metadatos por etiqueta (no por posición fija) encima del encabezado."""
    for fila in range(1, fila_encabezado):
        celdas = [ws.cell(row=fila, column=c).value for c in range(1, ws.max_column + 1)]
        textos = [(i, _norm(v)) for i, v in enumerate(celdas) if v is not None]
        for i, t in textos:
            resto = [v for v in celdas[i + 1:] if v is not None]
            valor = resto[0] if resto else None
            if t.startswith("fecha corte"):
                resultado.fecha_corte = str(valor or "")
            elif "ano al que se refiere" in t or "año" in t and "consulta" in t:
                try:
                    resultado.anio = int(valor)
                except (TypeError, ValueError):
                    pass
            elif t.startswith("tipo de documento"):
                resultado.tipo_documento = str(valor or "")
            elif t.startswith("identificacion:") or t == "identificacion":
                resultado.identificacion = str(valor or "")
            elif t.startswith("nombres / razon social") or t.startswith("nombres/razon social"):
                resultado.nombre = str(valor or "")
            elif t.startswith("fecha reporte") or "fecha reporte" in t:
                resultado.fecha_reporte = str(valor or "")
        # "Fecha   Reporte:" puede estar en cualquier columna con el valor al lado
        for i, t in textos:
            if "fecha" in t and "reporte" in t:
                resto = [v for v in celdas[i + 1:] if v is not None]
                if resto:
                    resultado.fecha_reporte = str(resto[0])


# --- reglas de clasificación -------------------------------------------------

# Detalles sin código R# que sí sabemos clasificar (texto normalizado → renglón)
_REGLAS_TEXTO = [
    # aportes obligatorios del trabajador (pensión/salud) = INCRNGO rentas de trabajo
    (re.compile(r"ingresos? no constitutivos? de renta"), 33,
     "INCRNGO de rentas de trabajo (aporte obligatorio del trabajador)"),
]

_NOTA_MULTIRENGLON = (
    "La exógena sugiere más de un renglón; se asignó al primero. "
    "Revise y reasigne en el resumen editable si corresponde."
)


def _clasificar(partida: PartidaExogena) -> None:
    """Aplica las reglas de asignación de renglón a una partida.

    Regla general: el primer R# mencionado es el renglón asignado; los demás
    quedan registrados en `renglones` como sugerencias secundarias que el
    usuario confirma en el resumen editable.
    """
    # Filas de referencia que la DIAN etiqueta con un renglón pero NO son un
    # valor a declarar (sumarían de más). Ej.: "ingreso laboral promedio de los
    # últimos seis meses" viene marcado R36, pero solo sirve para calcular el
    # tope de la renta exenta de cesantías — no es renta exenta en sí.
    det_n = _norm(partida.detalle or "")
    if "ingreso laboral promedio" in det_n:
        partida.excluida = True
        partida.nota = "Dato de referencia (ingreso promedio para el tope de cesantías): no suma."
        return

    # Cesantías consignadas al fondo (concepto 2276): la DIAN las etiqueta con
    # "Tope 1: Ingresos brutos | Tope 2: Patrimonio" (sin R#), así que quedaban sin
    # sumar. Son ingreso laboral del año → van a R32 (la dedup evita el doble conteo
    # con la fila "abonadas" que reporta el fondo).
    if "cesant" in det_n and "consignad" in det_n:
        partida.renglon_asignado = 32
        partida.renglones = [32]
        partida.nota = "Cesantías consignadas al fondo → ingresos laborales (R32)."
        return

    # Documentos soporte de adquisiciones (compras a no obligados a facturar):
    # NO se precargan en ninguna casilla — el contador decide. Si el cliente es
    # PN comerciante, esas compras ya deben ir DENTRO de las compras del módulo
    # (sumarlas aparte en R77 sería doble deducción); si es PN sin actividad,
    # no son deducibles (Art. 107 E.T.).
    if "documento" in det_n and "soporte" in det_n and "adquisic" in det_n:
        partida.renglon_asignado = None
        partida.renglones = [77]
        partida.nota = ("Documento soporte de adquisiciones: NO se precarga. "
                        "Comerciante: inclúyalo en las compras del módulo (CMV) si aplica. "
                        "PN sin actividad: no es deducible (Art. 107).")
        return

    uso = partida.uso_sugerido or ""
    partida.renglones = [int(m) for m in RE_RENGLON.findall(uso)]
    partida.topes = sorted({int(m) for m in RE_TOPE.findall(uso)})

    if partida.renglones:
        partida.renglon_asignado = partida.renglones[0]
        if len(set(partida.renglones)) > 1:
            partida.nota = _NOTA_MULTIRENGLON
            # caso conocido: R58 + R59 (rendimientos financieros: el ingreso va
            # completo a R58; el componente inflacionario a R59 según decreto anual)
            if set(partida.renglones) >= {58, 59}:
                partida.nota = (
                    "Rendimiento financiero: se suma completo a R58 y el componente "
                    "inflacionario (INCRNGO R59) se calcula con el % del archivo de "
                    "configuración (fijado por decreto anual)."
                )
        return

    uso_n = _norm(uso)
    for regla, renglon, nota in _REGLAS_TEXTO:
        if regla.search(uso_n):
            partida.renglon_asignado = renglon
            partida.nota = nota
            return

    if partida.topes:
        partida.nota = f"Solo informa Tope {', '.join(map(str, partida.topes))} (no suma a un renglón)."
    else:
        partida.nota = "Sin uso sugerido: partida informativa. Revise si debe declararla."


def _ajustar_beneficiario_economico(partida: PartidaExogena) -> Optional[str]:
    """Aplica la titularidad real (beneficiario económico) sobre el valor.

    La DIAN reporta el valor COMPLETO al titular principal aunque existan
    cotitulares u otros beneficiarios. Reglas:
      - 'Porcentaje de Participación: NN' < 100 → se ajusta el valor a la
        participación (se conserva el reportado en `valor_reportado`).
      - 'Número Propietarios' > 1 sin % informado → no se ajusta, pero se
        marca para revisión en el resumen editable.
      - menciones de cotitular/titular secundario/beneficiario → revisión.
    Devuelve una advertencia para el reporte, o None.
    """
    info = partida.info_adicional or ""
    m = RE_PARTICIPACION.search(info)
    if m:
        try:
            partida.participacion = float(m.group(1).replace(",", "."))
        except ValueError:
            partida.participacion = None
    m = RE_PROPIETARIOS.search(info)
    if m:
        partida.num_propietarios = int(m.group(1))

    if partida.participacion is not None and partida.participacion < 100:
        partida.valor_reportado = partida.valor
        partida.valor = round(partida.valor * partida.participacion / 100.0)
        partida.nota = (f"Ajustado al {partida.participacion:g}% de participación "
                        f"(reportado: {partida.valor_reportado:,.0f}). " + partida.nota).strip()
        return (f"Fila {partida.fila}: '{partida.detalle[:40]}' ajustada al "
                f"{partida.participacion:g}% de participación del beneficiario económico.")

    if (partida.num_propietarios or 1) > 1 and not partida.participacion:
        partida.nota = (f"Reportado 100% al titular principal pero hay "
                        f"{partida.num_propietarios} propietarios: declare solo su "
                        f"participación real (edite el valor). " + partida.nota).strip()
        return (f"Fila {partida.fila}: '{partida.detalle[:40]}' tiene "
                f"{partida.num_propietarios} propietarios y ningún % informado — "
                f"verifique el beneficiario económico y ajuste el valor.")

    if RE_COTITULAR.search(info) or RE_COTITULAR.search(partida.detalle or ""):
        partida.nota = ("Menciona cotitular/beneficiario: confirme qué parte le "
                        "corresponde como beneficiario económico real. " + partida.nota).strip()
        return (f"Fila {partida.fila}: '{partida.detalle[:40]}' menciona "
                f"cotitular/beneficiario — confirme la titularidad real.")
    return None


_TOPES_RESUMEN = {
    "tope 1": "ingresos",
    "tope 2": "patrimonio",
    "tope 3": "consumos_tc",
    "tope 4": "consignaciones",
    "tope 5": "compras",
}


def parsear_exogena(ruta, hoja: Optional[str] = None) -> ResultadoExogena:
    """Punto de entrada del parser."""
    ruta = Path(ruta)
    if not ruta.exists():
        raise ExogenaError(f"El archivo no existe: {ruta}")
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=False)
    except Exception as exc:
        raise ExogenaError(f"No se pudo abrir el archivo (¿corrupto o no es .xlsx?): {exc}") from exc

    if hoja:
        if hoja not in wb.sheetnames:
            raise ExogenaError(f"El libro no tiene la hoja '{hoja}'. Hojas: {wb.sheetnames}")
        ws = wb[hoja]
    elif "Reporte" in wb.sheetnames:
        ws = wb["Reporte"]
    else:
        ws = wb[wb.sheetnames[0]]

    resultado = ResultadoExogena(archivo=str(ruta))
    if ws.title != "Reporte":
        resultado.advertencias.append(
            f"La hoja procesada se llama '{ws.title}' (se esperaba 'Reporte')."
        )

    fila_enc, cols = _encontrar_encabezado(ws)
    if "uso" not in cols:
        raise ExogenaError(
            "El reporte no tiene la columna 'Uso declaración Sugerida', "
            "necesaria para clasificar las partidas."
        )
    _extraer_metadatos(ws, fila_enc, resultado)

    nits_informados = set()
    for fila in range(fila_enc + 1, ws.max_row + 1):
        def _celda(nombre):
            c = cols.get(nombre)
            return ws.cell(row=fila, column=c).value if c else None

        detalle = _celda("detalle")
        valor = _parse_valor(_celda("valor"))
        if detalle is None and valor is None:
            continue  # fila vacía intercalada

        detalle_n = _norm(detalle or "")
        # filas de resumen "Tope N - ..." al final del reporte
        es_resumen = False
        for prefijo, clave in _TOPES_RESUMEN.items():
            if detalle_n.startswith(prefijo):
                if valor is not None:
                    resultado.topes_dian[clave] = valor
                es_resumen = True
                break
        if es_resumen:
            continue

        if valor is None:
            resultado.advertencias.append(
                f"Fila {fila}: el valor '{_celda('valor')}' no es numérico; partida omitida "
                f"(detalle: {detalle})."
            )
            continue

        partida = PartidaExogena(
            fila=fila,
            informante_nit=str(_celda("informante_nit") or ""),
            informante_nombre=str(_celda("informante_nombre") or ""),
            informado_nit=str(_celda("informado_nit") or ""),
            informado_nombre=str(_celda("informado_nombre") or ""),
            detalle=str(detalle or ""),
            valor=valor,
            uso_sugerido=str(_celda("uso") or ""),
            info_adicional=str(_celda("info_adicional") or ""),
        )
        _clasificar(partida)
        aviso = _ajustar_beneficiario_economico(partida)
        if aviso:
            resultado.advertencias.append(aviso)
        resultado.partidas.append(partida)
        if partida.informado_nit:
            nits_informados.add(partida.informado_nit)

    if len(nits_informados) > 1:
        resultado.advertencias.append(
            f"El reporte contiene datos de más de un NIT informado: {sorted(nits_informados)}. "
            "Verifique que el archivo corresponda a un solo contribuyente."
        )
    if not resultado.partidas:
        resultado.advertencias.append("No se encontraron partidas en el reporte.")

    # Patrimonio y deudas duplicados: la DIAN consolida varios formatos y el
    # MISMO saldo (mismo tercero, mismo concepto, mismo valor) puede venir 2 o
    # 3 veces — sumarlo repetido infla el patrimonio. Se deja una sola fila
    # activa y las demás se excluyen con nota (el contador puede reactivarlas
    # si de verdad son bienes distintos con valores idénticos).
    vistos_pat = {}
    duplicadas = 0
    for pt in resultado.partidas:
        if pt.excluida or pt.renglon_asignado not in (29, 30):
            continue
        # Mismo tercero + mismo valor = el mismo bien, aunque el concepto venga
        # redactado distinto (la DIAN consolida formatos con textos diferentes).
        clave = (pt.renglon_asignado, pt.informante_nit, round(pt.valor or 0))
        if clave in vistos_pat:
            pt.excluida = True
            pt.nota = ("Mismo tercero y mismo valor ya contado en el patrimonio "
                       "(la exógena lo repite con otro concepto): se cuenta una "
                       "sola vez. Reactive la partida si son bienes distintos.")
            duplicadas += 1
        else:
            vistos_pat[clave] = pt
    if duplicadas:
        resultado.advertencias.append(
            f"Se excluyeron {duplicadas} fila(s) de patrimonio/deudas repetidas "
            "(mismo tercero, concepto y valor). Verifique en las partidas que no "
            "sean bienes distintos con valores idénticos.")

    # "Valor base del impuesto predial" vs "Valor avalúo catastral": el mismo
    # municipio reporta el mismo inmueble dos veces. Se declara por el MAYOR
    # valor (Art. 277 E.T.): queda la fila mayor (normalmente la base del
    # predial, que es el avalúo vigente) y se excluye la menor.
    predial_por_mun = {}
    for pt in resultado.partidas:
        if pt.excluida:
            continue
        det = _norm(pt.detalle)
        if "base del impuesto predial" in det:
            predial_por_mun.setdefault(pt.informante_nit, {}).setdefault("base", []).append(pt)
        elif "avaluo catastral" in det:
            predial_por_mun.setdefault(pt.informante_nit, {}).setdefault("avaluo", []).append(pt)
    for grupos in predial_por_mun.values():
        # Solo cuando el municipio reporta AMBOS conceptos (mismo(s) predio(s)
        # dos veces). Varios avalúos solos = varios predios: se respetan.
        if not grupos.get("base") or not grupos.get("avaluo"):
            continue
        tot_base = sum(f.valor or 0 for f in grupos["base"])
        tot_aval = sum(f.valor or 0 for f in grupos["avaluo"])
        perdedor = grupos["avaluo"] if tot_base >= tot_aval else grupos["base"]
        for f in perdedor:
            f.excluida = True
            f.nota = ("El municipio reporta el inmueble con dos conceptos (base "
                      "del predial y avalúo): se toma el de MAYOR valor "
                      "(Art. 277 E.T.) y esta fila se excluye.")

    # Saldo a favor del año anterior: la exógena lo trae como fila informativa
    # ("Total saldo a favor") sin renglón — va al R131 del 210.
    for pt in resultado.partidas:
        if not pt.excluida and pt.renglon_asignado is None \
                and (pt.valor or 0) > 0 \
                and ("total saldo a favor" in _norm(pt.detalle)
                     or "saldo a favor del ano anterior" in _norm(pt.detalle)
                     or "saldo a favor ano anterior" in _norm(pt.detalle)):
            pt.renglon_asignado = 131
            pt.nota = "Saldo a favor del año anterior (R131), reportado por la DIAN."

    # Aportes a salud del PENSIONADO: si quien los reporta es el mismo tercero
    # que paga la pensión (fondo/Colpensiones), el descuento va al INCRNGO de la
    # cédula de PENSIONES (R100), no al de rentas de trabajo (R33).
    pagadores_pension = {pt.informante_nit for pt in resultado.partidas
                         if pt.renglon_asignado == 99}
    for pt in resultado.partidas:
        if pt.excluida or pt.renglon_asignado != 33:
            continue
        det = _norm(pt.detalle)
        if "salud" not in det and "pension" not in det:
            continue
        inf = _norm(pt.informante_nombre)
        es_pagador = pt.informante_nit in pagadores_pension or             re.search(r"colpensiones|fondo de pensiones|pensiones y cesantias|"
                      r"porvenir|proteccion|colfondos|skandia", inf) is not None
        if es_pagador and pt.informante_nit in pagadores_pension:
            pt.renglon_asignado = 100
            pt.nota = ("Aporte a salud descontado por el pagador de la pensión: "
                       "va al INCRNGO de la cédula de pensiones (R100).")

    # Cesantías duplicadas: la MISMA cesantía puede venir reportada por el
    # EMPLEADOR ("consignadas al fondo") y por el FONDO ("abonadas en el
    # periodo") con idéntico valor y el mismo renglón — sumarla dos veces
    # infla el ingreso. Se deja una sola (se prefiere la del fondo) cuando
    # coinciden renglón y valor con informantes distintos.
    ces = [pt for pt in resultado.partidas
           if not pt.excluida and pt.renglon_asignado is not None
           and "cesant" in _norm(pt.detalle)
           and "intereses" not in _norm(pt.detalle)]
    grupos_ces = {}
    for pt in ces:
        grupos_ces.setdefault((pt.renglon_asignado, round(pt.valor or 0)), []).append(pt)
    dup_ces = 0
    for (_, _), filas_g in grupos_ces.items():
        if len(filas_g) < 2 or len({f.informante_nit for f in filas_g}) < 2:
            continue
        filas_g.sort(key=lambda f: 0 if "fondo" in _norm(f.informante_nombre) else 1)
        for f in filas_g[1:]:
            f.excluida = True
            f.nota = ("Cesantía reportada también por el fondo con el mismo valor: "
                      "se cuenta una sola vez.")
            dup_ces += 1
    if dup_ces:
        resultado.advertencias.append(
            f"Se excluyeron {dup_ces} fila(s) de cesantías duplicadas (mismo valor "
            "reportado por el empleador y por el fondo). Verifique en las partidas.")

    # Venta de activos FIJOS (Art. 300): la DIAN la manda a ganancia ocasional
    # (R112) por defecto, pero solo lo es si el activo se poseyó 2 años o más;
    # si menos, es RENTA NO LABORAL (R74). El tiempo de posesión no viene en la
    # exógena → se deja nota al contador para que lo verifique y reclasifique.
    ventas_af = [pt for pt in resultado.partidas if not pt.excluida
                 and "venta" in _norm(pt.detalle)
                 and "activo" in _norm(pt.detalle) and "fijo" in _norm(pt.detalle)]
    if ventas_af:
        total_af = sum(float(pt.valor or 0) for pt in ventas_af)
        con_fechas = False
        for pt in ventas_af:
            aviso = _nota_venta_activo(pt)     # usa fechas si la exógena las trae
            if "años de posesión" in aviso:
                con_fechas = True
            pt.nota = (aviso + " " + (pt.nota or "")).strip()
        base = (f"Detectamos venta de activos fijos por ${total_af:,.0f} (aplica a "
                "inmuebles y vehículos). Art. 300 E.T.: es GANANCIA OCASIONAL solo si el "
                "activo se poseyó 2 años o más; si MENOS de 2 años, la utilidad es RENTA "
                "NO LABORAL (cédula general).")
        if not con_fechas:
            base += (" La exógena no trae la fecha de adquisición: verifique el tiempo de "
                     "posesión y reclasifique (R112 GO ↔ R74 no laboral) si corresponde.")
        resultado.advertencias.append(base)

    # Deducción 1% factura electrónica (Art. 336 num. 5): si la DIAN reporta el
    # "monto susceptible de beneficio" (el que SÍ cumple los requisitos: pago
    # electrónico, no tomado como costo), esa es la base del R28 — no el total
    # bruto de facturas, que la exógena también trae y suele venir marcado R28.
    susceptible = [pt for pt in resultado.partidas
                   if "susceptible de beneficio" in _norm(pt.detalle)]
    if susceptible:
        for pt in susceptible:
            pt.renglon_asignado = 28
            pt.excluida = False
            pt.nota = "Base del 1% de factura electrónica (monto susceptible de beneficio)."
        for pt in resultado.partidas:
            if pt.renglon_asignado == 28 and pt not in susceptible                     and "suma valor total facturas" in _norm(pt.detalle):
                pt.excluida = True
                pt.nota = ("Total bruto de facturas electrónicas: informativo. La base "
                           "del 1% es el 'monto susceptible de beneficio'.")
    return resultado


# --- agregación de topes propios y obligación de declarar --------------------

def calcular_topes_propios(resultado: ResultadoExogena) -> Dict[str, float]:
    """Reagrega los topes desde las partidas para validarlos contra el resumen DIAN.

    Nota: el Tope 2 (patrimonio) de la DIAN toma el MAYOR entre la suma de
    variables del año y el patrimonio bruto declarado el año anterior.
    """
    mapa = {1: "ingresos", 2: "patrimonio", 3: "consumos_tc", 4: "consignaciones", 5: "compras"}
    tot: Dict[str, float] = {v: 0.0 for v in mapa.values()}
    patrimonio_anterior = 0.0
    for p in resultado.partidas_activas():
        if _norm(p.detalle).startswith("total patrimonio bruto declarado"):
            # valor de comparación, no aditivo: el tope toma el MAYOR
            patrimonio_anterior = max(patrimonio_anterior, p.valor)
            continue
        for t in p.topes:
            if t in mapa:
                tot[mapa[t]] += p.valor
    # compras: la DIAN usa la factura electrónica reportada (R28) si no hay filas
    # "Tope 5". Para el TOPE se toma el total bruto de facturas — incluso si la
    # fila quedó excluida del R28 (allí la base del 1% es el monto "susceptible
    # de beneficio", pero la obligación de declarar se mide con el consumo real).
    if tot["compras"] == 0.0:
        candidatos = [p.valor for p in resultado.partidas if 28 in p.renglones]
        candidatos += [p.valor for p in resultado.partidas
                       if "susceptible de beneficio" in _norm(p.detalle)]
        if candidatos:
            tot["compras"] = max(candidatos)
    tot["patrimonio"] = max(tot["patrimonio"], patrimonio_anterior)
    return tot


def evaluar_obligacion_declarar(topes: Dict[str, float], parametros) -> List[str]:
    """Devuelve la lista de razones por las que el contribuyente debe declarar."""
    umbral = parametros.topes_declarar_pesos()
    razones = []
    def _fmt(v): return f"${v:,.0f}"
    if topes.get("patrimonio", 0) > umbral["patrimonio_bruto"]:
        razones.append(
            f"Patrimonio bruto {_fmt(topes['patrimonio'])} supera "
            f"{parametros.topes_declarar_uvt['patrimonio_bruto']:,.0f} UVT ({_fmt(umbral['patrimonio_bruto'])})."
        )
    if topes.get("ingresos", 0) >= umbral["ingresos_brutos"]:
        razones.append(
            f"Ingresos brutos {_fmt(topes['ingresos'])} alcanzan "
            f"{parametros.topes_declarar_uvt['ingresos_brutos']:,.0f} UVT ({_fmt(umbral['ingresos_brutos'])})."
        )
    if topes.get("consumos_tc", 0) > umbral["consumos_tarjeta_credito"]:
        razones.append(
            f"Consumos con tarjeta {_fmt(topes['consumos_tc'])} superan "
            f"{parametros.topes_declarar_uvt['consumos_tarjeta_credito']:,.0f} UVT "
            f"({_fmt(umbral['consumos_tarjeta_credito'])})."
        )
    if topes.get("consignaciones", 0) > umbral["consignaciones_inversiones"]:
        razones.append(
            f"Consignaciones/inversiones {_fmt(topes['consignaciones'])} superan "
            f"{parametros.topes_declarar_uvt['consignaciones_inversiones']:,.0f} UVT "
            f"({_fmt(umbral['consignaciones_inversiones'])})."
        )
    if topes.get("compras", 0) > umbral["compras_consumos"]:
        razones.append(
            f"Compras y consumos {_fmt(topes['compras'])} superan "
            f"{parametros.topes_declarar_uvt['compras_consumos']:,.0f} UVT ({_fmt(umbral['compras_consumos'])})."
        )
    return razones
