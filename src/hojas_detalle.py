"""Llenado de las hojas de detalle de la plantilla ITGS.

Borra los datos de EJEMPLO que trae la plantilla (caso "Diana Zorani") y
escribe en su lugar los valores reales: las partidas de la exógena (con el
tercero que las reportó) y los datos capturados en la entrevista/app.

La hoja 'FORMULARIO 210' sigue siendo la autoritativa (valores estáticos del
motor); estas hojas quedan como soporte/anexo coherente con esos valores.
"""
import re
from typing import List, Optional

from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.hyperlink import Hyperlink

from .modelos import DatosDeclaracion, Liquidacion, PartidaExogena, ResultadoExogena
from .parametros import Parametros

# ----------------------------------------------------------------------
# identidad visual Tributando.co (hojas nuevas: Índice y Anexo Exógena)
# ----------------------------------------------------------------------
NAVY = "1E2432"
DORADO = "CDAB7E"
DORADO_OSCURO = "8A6D3B"       # dorado legible sobre blanco (texto)
GRIS_SUAVE = "F7F5F1"
BORDE_SUAVE = "E2DDD2"
# semáforo (fondo de fila / texto del estado)
VERDE_BG, VERDE_TX = "E8F5EE", "1F8A5F"
AMARILLO_BG, AMARILLO_TX = "FDF6E3", "9A6B00"
ROJO_BG, ROJO_TX = "FDECEA", "B3372F"

_FILL_NAVY = PatternFill("solid", fgColor=NAVY)
_FILL_GRIS = PatternFill("solid", fgColor=GRIS_SUAVE)
_BORDE_FINO = Border(bottom=Side(style="thin", color=BORDE_SUAVE))


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


# ----------------------------------------------------------------------
# utilidades de celda
# ----------------------------------------------------------------------

def _set(ws, coord: str, value) -> None:
    """Escribe aun si la celda pertenece a un rango combinado (usa el ancla)."""
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for rango in ws.merged_cells.ranges:
            if coord in rango:
                cell = ws.cell(row=rango.min_row, column=rango.min_col)
                break
    cell.value = value


def _limpiar_constantes(ws, rango: str) -> None:
    """Borra valores constantes (no fórmulas) en un rango A1:B9."""
    for fila in ws[rango]:
        for cell in fila:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v is not None and not (isinstance(v, str) and v.startswith("=")):
                cell.value = None


def _etiqueta(p: PartidaExogena, ancho: int = 84) -> str:
    """'Detalle — Informante (NIT)', garantizando que el informante quede visible.
    El NIT permite al contador cruzar contra la exógena sin ambigüedad."""
    inf = (p.informante_nombre or "").strip()
    if len(inf) > 28:
        inf = inf[:27] + "…"
    nit = (p.informante_nit or "").strip()
    if nit:
        inf = f"{inf} ({nit})" if inf else f"NIT {nit}"
    detalle = p.detalle.strip()
    if inf:
        max_detalle = max(20, ancho - len(inf) - 3)
        if len(detalle) > max_detalle:
            detalle = detalle[: max_detalle - 1] + "…"
        return f"{detalle} — {inf}"
    return detalle[:ancho]


def _comentar(ws, coord: str, p: PartidaExogena) -> None:
    """Comentario con la partida completa (sin recortes): al pasar el mouse el
    contador ve el concepto íntegro, el informante con NIT y la regla aplicada."""
    lineas = [p.detalle.strip()]
    if p.informante_nombre or p.informante_nit:
        lineas.append(f"Informante: {(p.informante_nombre or '').strip()} "
                      f"(NIT {(p.informante_nit or '—').strip()})")
    if p.valor_reportado is not None and abs(p.valor_reportado - p.valor) > 0.5:
        lineas.append(f"Valor reportado: ${p.valor_reportado:,.0f} → tomado: ${p.valor:,.0f}")
    if p.nota:
        lineas.append(f"Nota: {p.nota}")
    lineas.append(f"Fila {p.fila} de la exógena · Tributando.co")
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for rango in ws.merged_cells.ranges:
            if coord in rango:
                cell = ws.cell(row=rango.min_row, column=rango.min_col)
                break
    com = Comment("\n".join(lineas), "Tributando.co")
    com.width, com.height = 340, 150
    cell.comment = com


def _marcar_entrevista(ws, *coords) -> None:
    """Distingue en dorado itálica los valores que vienen de la entrevista/app
    (no de la exógena), para que se vea de dónde salió cada cifra."""
    for coord in coords:
        cell = ws[coord]
        if isinstance(cell, MergedCell):
            for rango in ws.merged_cells.ranges:
                if coord in rango:
                    cell = ws.cell(row=rango.min_row, column=rango.min_col)
                    break
        cell.font = Font(italic=True, color=DORADO_OSCURO)


def _escribir_lista(ws, partidas: List[PartidaExogena], filas: range,
                    col_desc: str, col_val: str, resto_etiqueta: str) -> float:
    """Escribe partidas una por fila; si no caben, agrega el resto agregado.
    Devuelve el total escrito."""
    filas = list(filas)
    total = 0.0
    individuales = partidas[: len(filas) - 1] if len(partidas) > len(filas) else partidas
    for i, p in enumerate(individuales):
        _set(ws, f"{col_desc}{filas[i]}", _etiqueta(p))
        _set(ws, f"{col_val}{filas[i]}", p.valor)
        _comentar(ws, f"{col_desc}{filas[i]}", p)
        total += p.valor
    resto = sum(p.valor for p in partidas[len(individuales):])
    if resto > 0:
        fila = filas[len(individuales)]
        _set(ws, f"{col_desc}{fila}", resto_etiqueta + " (detalle en Anexo Exógena)")
        _set(ws, f"{col_val}{fila}", resto)
        total += resto
    return total


def _match(p: PartidaExogena, *palabras) -> bool:
    d = p.detalle.lower()
    return any(w in d for w in palabras)


# ----------------------------------------------------------------------
# hojas
# ----------------------------------------------------------------------

def _hoja_deudas(wb, datos, exogena) -> None:
    if "Deudas" not in wb.sheetnames:
        return
    ws = wb["Deudas"]
    _limpiar_constantes(ws, "B2:C41")
    partidas = [p for p in _activas(exogena) if p.renglon_asignado == 30]
    total = _escribir_lista(ws, partidas, range(2, 40), "B", "C",
                            "Otras deudas reportadas")
    extra = datos.deudas - total
    if extra > 0.5:
        _set(ws, "B40", "Otras deudas declaradas (entrevista)")
        _set(ws, "C40", extra)
        _marcar_entrevista(ws, "B40", "C40")


def _hoja_pat_bruto(wb, datos, exogena) -> None:
    if "Pat bruto" not in wb.sheetnames:
        return
    ws = wb["Pat bruto"]
    # limpiar ejemplos: financieras, cxc, acciones, activos fijos, opción 2
    _limpiar_constantes(ws, "B5:D28")
    _limpiar_constantes(ws, "B33:D57")
    _limpiar_constantes(ws, "B63:E78")
    _limpiar_constantes(ws, "B84:E100")
    for f in range(84, 101):        # avalúos catastrales de ejemplo (col H)
        _limpiar_constantes(ws, f"H{f}:H{f}")
    _limpiar_constantes(ws, "B106:F116")
    for f in range(84, 101):        # flag de reajuste: dejar en 'No'
        _set(ws, f"C{f}", "No")
    for f in range(63, 79):
        _set(ws, f"C{f}", "No")

    partidas = [p for p in _activas(exogena) if p.renglon_asignado == 29]
    avaluos = [p for p in partidas if _match(p, "avalúo", "avaluo", "catastral")]
    cxc = [p for p in partidas if _match(p, "cuenta por cobrar", "cuentas por cobrar")]
    financieras = [p for p in partidas if p not in avaluos and p not in cxc]

    total = _escribir_lista(ws, financieras, range(5, 29), "B", "D",
                            "Otros saldos financieros")
    total += _escribir_lista(ws, cxc, range(33, 58), "B", "D",
                             "Otras cuentas por cobrar")
    # inmuebles: el valor va como avalúo (col H); I toma el mayor vs costo
    filas_af = list(range(84, 101))
    for i, p in enumerate(avaluos[: len(filas_af)]):
        extra_info = (p.info_adicional or "").strip()[:40]
        _set(ws, f"B{filas_af[i]}", _etiqueta(p) + (f" ({extra_info})" if extra_info else ""))
        _set(ws, f"H{filas_af[i]}", p.valor)
        total += p.valor
    # activos adicionales de la entrevista/app
    extra = datos.patrimonio_bruto - total
    if extra > 0.5:
        fila = filas_af[min(len(avaluos), len(filas_af) - 1)]
        _set(ws, f"B{fila}", "Otros activos declarados (vehículos, muebles, efectivo — entrevista)")
        _set(ws, f"D{fila}", extra)
        _marcar_entrevista(ws, f"B{fila}", f"D{fila}")


def _hoja_retefuente(wb, datos, exogena) -> None:
    if "retefuente" not in wb.sheetnames:
        return
    ws = wb["retefuente"]
    _limpiar_constantes(ws, "C3:D36")
    partidas = [p for p in _activas(exogena) if p.renglon_asignado == 132]
    total = _escribir_lista(ws, partidas, range(3, 35), "C", "D",
                            "Otras retenciones reportadas")
    extra = datos.retenciones - total
    if extra > 0.5:
        _set(ws, "C35", "Otras retenciones (entrevista)")
        _set(ws, "D35", extra)
        _marcar_entrevista(ws, "C35", "D35")


def _hoja_anticipo(wb, datos) -> None:
    if "anticipo" not in wb.sheetnames:
        return
    ws = wb["anticipo"]
    _set(ws, "E3", datos.saldo_favor_anterior)
    _set(ws, "E4", datos.anticipo_anterior)
    _set(ws, "E5", datos.impuesto_neto_anio_anterior)
    _set(ws, "E6", min(max(datos.numero_anio_declaracion, 1), 3))


def _hoja_dtos_tribut(wb, datos) -> None:
    if "dtos tribut" not in wb.sheetnames:
        return
    ws = wb["dtos tribut"]
    _limpiar_constantes(ws, "B3:D5")
    _limpiar_constantes(ws, "B8:C9")
    _limpiar_constantes(ws, "D21:D23")
    if datos.descuento_impuestos_exterior:
        _set(ws, "B3", "Impuestos pagados en el exterior")
        _set(ws, "D3", datos.descuento_impuestos_exterior)
    if datos.descuento_donaciones:
        _set(ws, "B8", "Donaciones certificadas (entrevista)")
        # la hoja calcula D8 = C8 × 25%; se registra la base equivalente
        _set(ws, "C8", round(datos.descuento_donaciones / 0.25))
    if datos.descuento_go_exterior:
        _set(ws, "D21", datos.descuento_go_exterior)


def _hoja_trabajo(wb, datos, liq, exogena) -> None:
    nombre = "R.trabajo y honorarios"
    if nombre not in wb.sheetnames:
        return
    ws = wb[nombre]
    t, h = datos.trabajo, datos.honorarios

    # ---- ingresos (G4:G19 trabajo, J9 honorarios) ----
    _limpiar_constantes(ws, "G4:G19")
    _limpiar_constantes(ws, "J9:J19")
    partidas = [p for p in _activas(exogena) if p.renglon_asignado == 32]
    salarios = sum(p.valor for p in partidas if _match(p, "pagos por salarios"))
    cesantias_cons = sum(p.valor for p in partidas
                         if _match(p, "cesantías consignadas", "cesantias consignadas",
                                   "cesantías abonadas", "cesantias abonadas"))
    resto = t.ingresos_brutos - salarios - cesantias_cons
    _set(ws, "G4", salarios)
    _set(ws, "G6", cesantias_cons)   # reemplaza la fórmula de ejemplo =G4/12
    _set(ws, "G12", max(0.0, resto))
    _set(ws, "J9", h.ingresos_brutos)

    # ---- INCRNGO (G21:G32) ----
    _limpiar_constantes(ws, "G21:G32")
    _limpiar_constantes(ws, "J22:K32")
    incr = [p for p in _activas(exogena) if p.renglon_asignado == 33]
    pension = sum(p.valor for p in incr if _match(p, "pensión", "pension"))
    salud = sum(p.valor for p in incr if _match(p, "salud"))
    _set(ws, "G21", pension + max(0.0, t.incrngo - pension - salud))
    _set(ws, "G23", salud)
    _set(ws, "J25", h.incrngo)

    # ---- costos honorarios ----
    _limpiar_constantes(ws, "J34:K47")
    _set(ws, "J36", h.costos_deducciones)

    # ---- rentas exentas y deducciones (valores del motor) ----
    _limpiar_constantes(ws, "F50:F67")
    _limpiar_constantes(ws, "J50:J67")
    _set(ws, "F50", t.rentas_exentas_afc_fvp)
    _set(ws, "G50", t.rentas_exentas_afc_fvp)          # reemplaza fórmula Sheet1
    _set(ws, "F52", t.otras_rentas_exentas)
    _set(ws, "G52", t.otras_rentas_exentas)
    exenta_25 = liq.r(36) - round(t.otras_rentas_exentas)
    _set(ws, "G62", max(0.0, exenta_25))               # renta exenta 25% calculada
    _set(ws, "F63", t.intereses_vivienda)
    _set(ws, "F64", "No")                              # dependientes van en R139
    _set(ws, "G64", 0)
    _set(ws, "F65", t.otras_deducciones)
    _set(ws, "G65", t.otras_deducciones)               # sin recorte de fórmula ejemplo
    _set(ws, "K50", h.rentas_exentas_afc_fvp)
    _set(ws, "J63", h.intereses_vivienda)
    _set(ws, "J64", "No")
    _set(ws, "K64", 0)


def _hoja_capital(wb, datos, exogena, p: Parametros) -> None:
    if "R.capital" not in wb.sheetnames:
        return
    ws = wb["R.capital"]
    c = datos.capital
    _limpiar_constantes(ws, "C4:E11")
    _limpiar_constantes(ws, "C14:E18")
    _limpiar_constantes(ws, "C22:E27")
    _limpiar_constantes(ws, "C30:E35")
    _limpiar_constantes(ws, "E40:F40")
    _limpiar_constantes(ws, "F37:F39")
    _limpiar_constantes(ws, "C46:F51")
    _limpiar_constantes(ws, "E54:E57")

    partidas = [q for q in _activas(exogena) if q.renglon_asignado == 58]
    rendimientos = [q for q in partidas if 59 in q.renglones]
    otros = [q for q in partidas if 59 not in q.renglones]
    arriendos = [q for q in otros if _match(q, "arrendamiento", "arriendo")]
    otros = [q for q in otros if q not in arriendos]

    total = _escribir_lista(ws, rendimientos, range(4, 12), "C", "E",
                            "Otros rendimientos financieros")
    total += _escribir_lista(ws, arriendos, range(22, 28), "C", "E", "Otros arriendos")
    total += _escribir_lista(ws, otros, range(14, 19), "C", "E",
                             "Otros ingresos de capital")
    extra = c.ingresos_brutos - total
    if extra > 0.5:
        _set(ws, "C18", "Otros ingresos de capital (entrevista)")
        _set(ws, "E18", extra)
        _marcar_entrevista(ws, "C18", "E18")
    # el componente inflacionario queda como fórmula F41 = F3 × % (config)
    _set(ws, "F41", f"=F3*{p.componente_inflacionario*100:.2f}%")
    if c.costos_deducciones:
        _set(ws, "C46", "Costos y gastos de rentas de capital (entrevista)")
        _set(ws, "F46", c.costos_deducciones)
        _marcar_entrevista(ws, "C46", "F46")
    _set(ws, "E55", c.intereses_vivienda)


def _hoja_no_laboral(wb, datos, exogena) -> None:
    nombre = "R.no laboral y R gravables"
    if nombre not in wb.sheetnames:
        return
    ws = wb[nombre]
    nl = datos.no_laboral
    _limpiar_constantes(ws, "C3:F16")
    _limpiar_constantes(ws, "C18:F22")
    _limpiar_constantes(ws, "F25:F27")
    _limpiar_constantes(ws, "C39:F56")
    _limpiar_constantes(ws, "E64:E71")
    _limpiar_constantes(ws, "F83:F85")

    partidas = [q for q in _activas(exogena) if q.renglon_asignado == 74]
    total = _escribir_lista(ws, partidas, range(3, 16), "C", "F",
                            "Otros ingresos no laborales")
    extra = nl.ingresos_brutos - total
    if extra > 0.5:
        _set(ws, "C16", "Otros ingresos no laborales (entrevista)")
        _set(ws, "F16", extra)
        _marcar_entrevista(ws, "C16", "F16")
    if nl.devoluciones:
        _set(ws, "C18", "Devoluciones, rebajas y descuentos")
        _set(ws, "F18", nl.devoluciones)
    if nl.incrngo:
        _set(ws, "C25", "INCRNGO rentas no laborales")
        _set(ws, "F25", nl.incrngo)
    if nl.costos_deducciones:
        _set(ws, "C39", "Costos y gastos de rentas no laborales")
        _set(ws, "F39", nl.costos_deducciones)
    if datos.rentas_gravables:
        _set(ws, "F83", datos.rentas_gravables)


def _hoja_pensiones(wb, datos) -> None:
    if "C.pensiones" not in wb.sheetnames:
        return
    ws = wb["C.pensiones"]
    _limpiar_constantes(ws, "E3:E8")
    _limpiar_constantes(ws, "E11:E15")
    _limpiar_constantes(ws, "E18:E24")
    if datos.pension_ingresos:
        _set(ws, "C3", "Ingresos por pensiones (fuente nacional)")
        _set(ws, "E3", datos.pension_ingresos)
    if datos.pension_incrngo:
        _set(ws, "E19", datos.pension_incrngo)
    # la exención de la plantilla (E26/F26) se reemplaza por el valor declarado
    _set(ws, "E26", min(datos.pension_exenta, datos.pension_ingresos))
    _set(ws, "F26", min(datos.pension_exenta, datos.pension_ingresos))


def _hoja_dividendos(wb, datos) -> None:
    if "C divid." not in wb.sheetnames:
        return
    ws = wb["C divid."]
    for rango in ("E3:E9", "E11:E15", "E17:E23", "E25:E31", "E33:E39", "E41:E47"):
        _limpiar_constantes(ws, rango)
    if datos.dividendos_2016_anteriores:
        _set(ws, "E4", datos.dividendos_2016_anteriores)
    if datos.dividendos_2016_incrngo:
        _set(ws, "E12", datos.dividendos_2016_incrngo)   # reemplaza fórmula =E4
    if datos.dividendos_sub1:
        _set(ws, "E18", datos.dividendos_sub1)
    if datos.dividendos_sub2:
        _set(ws, "E26", datos.dividendos_sub2)
    if datos.dividendos_exterior:
        _set(ws, "E33", datos.dividendos_exterior)
    if datos.dividendos_exterior_exentos:
        _set(ws, "E41", datos.dividendos_exterior_exentos)


def _hoja_gocas(wb, datos, liq: Liquidacion, p: Parametros) -> None:
    if "G.OCAS" not in wb.sheetnames:
        return
    ws = wb["G.OCAS"]
    # limpiar la venta y herencias de ejemplo
    _limpiar_constantes(ws, "C3:D3")
    _limpiar_constantes(ws, "F3:F6")
    _limpiar_constantes(ws, "F8:F10")
    _limpiar_constantes(ws, "F17:F20")
    _limpiar_constantes(ws, "F24:F24")
    _limpiar_constantes(ws, "D46:D66")
    _set(ws, "E89", p.uvt)

    # Se recorren las partidas efectivas para que funcione igual con entrada
    # plana (campos go_*) que con ganancias tipificadas.
    ingresos_15 = costos_15 = loterias = 0.0
    descripciones = []
    for go in datos.go_partidas_efectivas():
        if p.go_tipo(go.tipo).get("loteria"):
            loterias += go.ingreso
        else:
            ingresos_15 += go.ingreso
            costos_15 += go.costo_fiscal
            if go.descripcion:
                descripciones.append(go.descripcion)

    if ingresos_15:
        _set(ws, "C3", "; ".join(descripciones)
             or "Ganancias ocasionales declaradas (detalle en entrevista)")
        _set(ws, "D3", ingresos_15)
        _set(ws, "G3", costos_15)            # reemplaza la fórmula de costo fiscal
    if loterias:
        _set(ws, "D63", loterias)
    if liq.r(114):
        _set(ws, "D83", liq.r(114))          # reemplaza el cálculo de ejemplo


def _estado_partida(p: PartidaExogena) -> str:
    if p.excluida:
        return "Excluida"
    if p.valor_reportado is not None and abs(p.valor_reportado - p.valor) > 0.5:
        return "Ajustada"
    return "Incluida"


def _hoja_anexo(wb, datos, liq: Liquidacion, exogena: Optional[ResultadoExogena]) -> None:
    """Anexo Exógena: el cruce completo exógena → Formulario 210, partida por
    partida y SIN recortes. Semáforo por estado, filtros y encabezado fijo.
    Es el papel de trabajo de auditoría del contador."""
    if exogena is None or not exogena.partidas:
        return
    nombre = "Anexo Exógena"
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False

    # ---- título ----
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "  CRUCE EXÓGENA → FORMULARIO 210 · papel de trabajo"
    c.fill = _FILL_NAVY
    c.font = Font(bold=True, size=13, color=DORADO)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (f"  {exogena.nombre or ''} · NIT {exogena.identificacion or '—'} · "
               f"año {exogena.anio or ''} · generado por Tributando.co")
    c.fill = _FILL_NAVY
    c.font = Font(size=9, color="FFFFFF")
    ws.row_dimensions[2].height = 16

    # ---- encabezados ----
    cabeceras = ["Fila", "Concepto reportado", "Informante", "NIT informante",
                 "Valor reportado", "Valor tomado", "Renglón 210", "Estado", "Nota"]
    for j, texto in enumerate(cabeceras, start=1):
        c = ws.cell(row=4, column=j, value=texto)
        c.fill = _FILL_GRIS
        c.font = Font(bold=True, size=9, color=NAVY)
        c.border = Border(bottom=Side(style="medium", color=DORADO))
        c.alignment = Alignment(vertical="center", wrap_text=True)

    # ---- partidas (todas: incluidas, ajustadas y excluidas) ----
    colores = {"Incluida": (None, VERDE_TX), "Ajustada": (AMARILLO_BG, AMARILLO_TX),
               "Excluida": (ROJO_BG, ROJO_TX)}
    fila = 5
    total_tomado = 0.0
    for p in sorted(exogena.partidas, key=lambda q: q.fila):
        estado = _estado_partida(p)
        bg, tx = colores[estado]
        reportado = p.valor_reportado if p.valor_reportado is not None else p.valor
        tomado = 0.0 if p.excluida else p.valor
        total_tomado += tomado
        valores = [p.fila, p.detalle.strip(), (p.informante_nombre or "").strip(),
                   (p.informante_nit or "").strip(), reportado, tomado,
                   p.renglon_asignado or "—", estado, (p.nota or "").strip()]
        for j, v in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=j, value=v)
            c.border = _BORDE_FINO
            c.font = Font(size=9, color=NAVY)
            if bg:
                c.fill = _fill(bg)
            if j in (5, 6):
                c.number_format = "#,##0"
            if j == 2 or j == 9:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=fila, column=8).font = Font(size=9, bold=True, color=tx)
        fila += 1

    # ---- total ----
    c = ws.cell(row=fila, column=2, value="TOTAL TOMADO EN LA DECLARACIÓN")
    c.font = Font(bold=True, size=9, color=NAVY)
    c = ws.cell(row=fila, column=6, value=total_tomado)
    c.font = Font(bold=True, size=10, color=DORADO_OSCURO)
    c.number_format = "#,##0"
    for j in range(1, 10):
        ws.cell(row=fila, column=j).border = Border(
            top=Side(style="medium", color=DORADO))

    # ---- advertencias del análisis ----
    advertencias = list(exogena.advertencias or []) + list(liq.advertencias or [])
    if advertencias:
        fila += 2
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
        c = ws.cell(row=fila, column=1, value="  ⚠ ADVERTENCIAS DEL ANÁLISIS — revisar con criterio profesional")
        c.fill = _FILL_NAVY
        c.font = Font(bold=True, size=10, color=DORADO)
        ws.row_dimensions[fila].height = 20
        for adv in advertencias:
            fila += 1
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
            c = ws.cell(row=fila, column=1, value=f"• {adv}")
            c.font = Font(size=9, color=AMARILLO_TX)
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- interactividad: filtros + encabezado fijo + anchos ----
    ws.auto_filter.ref = f"A4:I{4 + len(exogena.partidas)}"
    ws.freeze_panes = "A5"
    anchos = {"A": 6, "B": 52, "C": 26, "D": 15, "E": 14, "F": 14,
              "G": 11, "H": 11, "I": 42}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho


def _hoja_indice(wb) -> None:
    """Portada Índice: navegación con un clic a cada hoja, con la marca."""
    nombre = "Índice"
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre, 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "  Tributando.co"
    c.fill = _FILL_NAVY
    c.font = Font(bold=True, size=16, color=DORADO)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 34
    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = "  Papeles de trabajo — Declaración de Renta · Formulario 210"
    c.fill = _FILL_NAVY
    c.font = Font(size=10, color="FFFFFF")
    ws.row_dimensions[3].height = 18

    secciones = [
        ("FORMULARIO 210", "🧾  Formulario 210 (autoritativo)"),
        ("Anexo Exógena", "📋  Anexo Exógena — cruce completo y advertencias"),
        ("Pat bruto", "🏠  Patrimonio bruto"),
        ("Deudas", "💳  Deudas"),
        ("R.trabajo y honorarios", "💼  Rentas de trabajo y honorarios"),
        ("R.capital", "🏦  Rentas de capital"),
        ("R.no laboral y R gravables", "📦  Rentas no laborales y gravables"),
        ("C.pensiones", "👴  Cédula de pensiones"),
        ("C divid.", "📈  Dividendos"),
        ("G.OCAS", "🎯  Ganancias ocasionales"),
        ("Dependientes ", "👨‍👩‍👧  Dependientes"),
        ("retefuente", "🧮  Retenciones en la fuente"),
        ("anticipo", "⏭  Anticipo de renta"),
        ("dtos tribut", "🎫  Descuentos tributarios"),
    ]
    fila = 5
    for i, (hoja, etiqueta) in enumerate(secciones):
        if hoja not in wb.sheetnames:
            continue
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=5)
        c = ws.cell(row=fila, column=2, value=f"  {etiqueta}")
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{hoja}'!A1")
        c.font = Font(size=11, color=NAVY, underline="single")
        if i % 2 == 0:
            for j in range(2, 6):
                ws.cell(row=fila, column=j).fill = _FILL_GRIS
        ws.row_dimensions[fila].height = 22
        fila += 1

    fila += 1
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=5)
    c = ws.cell(row=fila, column=2,
                value="  Clic en una sección para ir directo. Los valores en dorado itálica "
                      "provienen de la entrevista/app; el resto, de la exógena DIAN.")
    c.font = Font(size=8, italic=True, color=DORADO_OSCURO)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 3
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 16
    wb.active = 0


def _actualizar_uvt(wb, p: Parametros) -> None:
    """La plantilla trae UVT de años anteriores en varias celdas."""
    if "R.trabajo y honorarios" in wb.sheetnames:
        wb["R.trabajo y honorarios"]["H74"] = p.uvt
    if "impto renta" in wb.sheetnames:
        wb["impto renta"]["E2"] = p.uvt
    if "Dependientes " in wb.sheetnames:
        ws = wb["Dependientes "]
        _set(ws, "C4", round(p.dependientes_uvt * p.uvt))
        _set(ws, "D12", round(p.dependientes_uvt * p.uvt))


def _activas(exogena: Optional[ResultadoExogena]) -> List[PartidaExogena]:
    return exogena.partidas_activas() if exogena else []


def llenar_hojas_detalle(wb, datos: DatosDeclaracion, liq: Liquidacion,
                         exogena: Optional[ResultadoExogena],
                         parametros: Optional[Parametros] = None) -> None:
    """Punto de entrada: llena todas las hojas de soporte y borra ejemplos."""
    p = parametros or Parametros.cargar(2025)
    _hoja_deudas(wb, datos, exogena)
    _hoja_pat_bruto(wb, datos, exogena)
    _hoja_retefuente(wb, datos, exogena)
    _hoja_anticipo(wb, datos)
    _hoja_dtos_tribut(wb, datos)
    _hoja_trabajo(wb, datos, liq, exogena)
    _hoja_capital(wb, datos, exogena, p)
    _hoja_no_laboral(wb, datos, exogena)
    _hoja_pensiones(wb, datos)
    _hoja_dividendos(wb, datos)
    _hoja_gocas(wb, datos, liq, p)
    _actualizar_uvt(wb, p)
    _hoja_anexo(wb, datos, liq, exogena)
    _hoja_indice(wb)                    # al final, para enlazar el anexo
