"""Resumen ejecutivo en PDF para entregar al cliente.

Genera un documento de 1-2 páginas con: datos del declarante, obligación de
declarar, composición por cédulas, liquidación del impuesto, resultado
(saldo a pagar / a favor) y advertencias, con el aviso de borrador.
"""
from datetime import date
from pathlib import Path
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (HRFlowable, Image, Paragraph, SimpleDocTemplate,
                                Spacer, Table, TableStyle)

from .modelos import DatosDeclaracion, Liquidacion, ResultadoExogena
from .parametros import Parametros

AZUL = colors.HexColor("#1E2432")   # azul Tributando
GRIS = colors.HexColor("#f6f4ee")   # crema Tributando
ORO = colors.HexColor("#E0C584")    # oro Tributando
ROJO = colors.HexColor("#b3372f")
VERDE = colors.HexColor("#1e7d43")


def _fmt(v: float) -> str:
    return f"${v:,.0f}".replace(",", ".")


def _estilos():
    base = getSampleStyleSheet()
    return {
        "titulo": ParagraphStyle("t", parent=base["Title"], fontSize=15,
                                 textColor=AZUL, spaceAfter=2),
        "sub": ParagraphStyle("s", parent=base["Normal"], fontSize=8.5,
                              textColor=colors.HexColor("#5a6b7f"),
                              alignment=TA_CENTER, spaceAfter=10),
        "h2": ParagraphStyle("h2", parent=base["Heading2"], fontSize=11,
                             textColor=AZUL, spaceBefore=10, spaceAfter=4),
        "normal": ParagraphStyle("n", parent=base["Normal"], fontSize=9, leading=12),
        "peq": ParagraphStyle("p", parent=base["Normal"], fontSize=7.5,
                              textColor=colors.HexColor("#5a6b7f"), leading=9.5),
        "alerta": ParagraphStyle("a", parent=base["Normal"], fontSize=8.5,
                                 textColor=ROJO, leading=11),
        "kpi": ParagraphStyle("k", parent=base["Normal"], fontSize=13,
                              alignment=TA_RIGHT, leading=15),
    }


def _tabla(filas, anchos, negrilla_ultima=False, resaltar=None):
    t = Table(filas, colWidths=anchos)
    estilo = [
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, 0), ORO),
        ("BACKGROUND", (0, 0), (-1, 0), AZUL),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2dccc")),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    if negrilla_ultima:
        estilo += [("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                   ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f0ebdd"))]
    if resaltar is not None:
        estilo += [("TEXTCOLOR", (1, resaltar), (1, resaltar),
                    ROJO if "pagar" in str(filas[resaltar][0]).lower() else VERDE)]
    t.setStyle(TableStyle(estilo))
    return t


def _norm(t: str) -> str:
    import unicodedata
    t = unicodedata.normalize("NFD", (t or "").lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def _top_terceros(exogena, renglones, n=7):
    """[(informante, detalle corto, valor)] de las partidas activas de esos renglones."""
    acum = {}
    for pt in exogena.partidas_activas():
        if pt.renglon_asignado in renglones:
            k = (pt.informante_nombre[:38], pt.detalle[:52])
            acum[k] = acum.get(k, 0) + pt.valor
    filas = sorted(acum.items(), key=lambda kv: -kv[1])[:n]
    return [(a, b, v) for (a, b), v in filas]


def _retenciones_por_tercero(exogena, n=8):
    acum = {}
    for pt in exogena.partidas_activas():
        if "retencion" in _norm(pt.detalle) and pt.renglon_asignado == 132:
            acum[pt.informante_nombre[:44]] = acum.get(pt.informante_nombre[:44], 0) + pt.valor
    return sorted(acum.items(), key=lambda kv: -kv[1])[:n]


def _partidas_renglon(exogena, renglon, n=12):
    """[(informante, detalle corto, valor)] de las partidas activas de un renglón."""
    filas = [(pt.informante_nombre[:36], pt.detalle[:50], pt.valor)
             for pt in exogena.partidas_activas() if pt.renglon_asignado == renglon]
    filas.sort(key=lambda f: -f[2])
    return filas[:n]


def _patrimonio_anterior(exogena):
    for pt in exogena.partidas:
        if "patrimonio bruto declarado en el ano anterior" in _norm(pt.detalle):
            return pt.valor or 0
    return 0


def generar_resumen_pdf(
    ruta: Path,
    datos: DatosDeclaracion,
    liq: Liquidacion,
    p: Parametros,
    exogena: Optional[ResultadoExogena] = None,
    razones_obligado=None,
    preparado_por: str = "",
    fecha_lim=None,
    observaciones: str = "",
    liq_base=None,
    logo_bytes: Optional[bytes] = None,
) -> Path:
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    st = _estilos()
    R = liq.r
    con = datos.contribuyente
    nombre = " ".join(x for x in (con.primer_nombre, con.otros_nombres,
                                  con.primer_apellido, con.segundo_apellido) if x) \
             or (exogena.nombre if exogena else "")

    doc = SimpleDocTemplate(str(ruta), pagesize=letter,
                            topMargin=16 * mm, bottomMargin=14 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            title=f"Resumen ejecutivo renta AG {p.anio_gravable}")
    e = []

    # ---------------- encabezado (con el logo de la firma del contador) ----------------
    titulo_flow = [Paragraph("Resumen Ejecutivo — Declaración de Renta", st["titulo"]),
                   Paragraph(
        f"Formulario 210 · Personas Naturales Residentes · Año Gravable {p.anio_gravable} · "
        f"UVT {_fmt(p.uvt)} · Preparado el {date.today().strftime('%d/%m/%Y')}"
        + (f" · {preparado_por}" if preparado_por else ""), st["sub"])]
    if logo_bytes:
        try:
            import io as _io
            from reportlab.lib.utils import ImageReader
            img = ImageReader(_io.BytesIO(logo_bytes))
            iw, ih = img.getSize()
            alto = 22 * mm
            ancho = min(alto * iw / ih, 70 * mm)
            if ancho == 70 * mm:                 # logos muy anchos: manda el ancho
                alto = min(alto, 70 * mm * ih / iw)
            logo = Image(_io.BytesIO(logo_bytes), width=ancho, height=alto)
            cab = Table([[titulo_flow, logo]], colWidths=[None, ancho + 4])
            cab.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                     ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                                     ("LEFTPADDING", (0, 0), (-1, -1), 0),
                                     ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
            e.append(cab)
        except Exception:
            e.extend(titulo_flow)          # logo dañado: se sigue sin él
    else:
        e.extend(titulo_flow)
    e.append(HRFlowable(width="100%", color=AZUL, thickness=1.2))

    # ---------------- declarante ----------------
    e.append(Paragraph("1. Datos del declarante", st["h2"]))
    e.append(_tabla([
        ["Contribuyente", "Identificación", "DV", "Actividad económica", "Dependientes"],
        [nombre or "—", con.nit or "—", con.dv or "—",
         con.actividad_economica or "—", str(datos.dependientes)],
    ], [None, 90, 30, 95, 70]))

    # ---------------- obligación ----------------
    if razones_obligado:
        e.append(Paragraph("2. Obligación de declarar", st["h2"]))
        e.append(Paragraph(
            "Según la información exógena reportada por terceros a la DIAN, el "
            "contribuyente <b>está obligado a declarar</b> por:", st["normal"]))
        for r in razones_obligado:
            e.append(Paragraph(f"• {r}", st["normal"]))

    # ---------------- calendario ----------------
    if fecha_lim:
        dias = (fecha_lim - date.today()).days
        e.append(Paragraph("2b. Plazo para declarar", st["h2"]))
        e.append(Paragraph(
            f"Fecha límite según los dos últimos dígitos del documento: "
            f"<b>{fecha_lim.strftime('%d/%m/%Y')}</b> "
            f"({'quedan <b>%d días</b>' % dias if dias >= 0 else '<b>VENCIDA hace %d días</b>' % -dias}). "
            f"Presentar después de la fecha causa sanción de extemporaneidad "
            f"(mínima {_fmt(10 * p.uvt_presentacion)} = 10 UVT de {p.anio_gravable + 1}) "
            f"más intereses de mora.", st["normal"]))

    # ---------------- patrimonio ----------------
    e.append(Paragraph("3. Patrimonio detallado", st["h2"]))
    bienes = _partidas_renglon(exogena, 29) if exogena else []
    if bienes:
        suma_rep = sum(v for *_, v in bienes)
        filas_b = [["Bien / derecho (reportado por)", "Detalle", "Valor"]]
        filas_b += [[a, b, _fmt(v)] for a, b, v in bienes]
        if R(29) - suma_rep > 0.5:
            filas_b.append(["Otros activos no reportados en exógena",
                            "(vehículos, inmuebles, efectivo, ajustes)", _fmt(R(29) - suma_rep)])
        filas_b.append(["TOTAL PATRIMONIO BRUTO (R29)", "", _fmt(R(29))])
        e.append(_tabla(filas_b, [150, None, 90], negrilla_ultima=True))
    deudas = _partidas_renglon(exogena, 30) if exogena else []
    if deudas or R(30):
        e.append(Spacer(1, 4))
        suma_d = sum(v for *_, v in deudas)
        filas_d = [["Deuda (reportada por)", "Detalle", "Valor"]]
        filas_d += [[a, b, _fmt(v)] for a, b, v in deudas]
        if R(30) - suma_d > 0.5:
            filas_d.append(["Otras deudas no reportadas en exógena", "", _fmt(R(30) - suma_d)])
        filas_d.append(["TOTAL DEUDAS (R30)", "", _fmt(R(30))])
        e.append(_tabla(filas_d, [150, None, 90], negrilla_ultima=True))
    e.append(Spacer(1, 4))
    e.append(_tabla([
        ["Concepto", "Renglón", "Valor"],
        ["Patrimonio bruto", "29", _fmt(R(29))],
        ["(−) Deudas", "30", _fmt(R(30))],
        ["Patrimonio líquido", "31", _fmt(R(31))],
    ], [None, 60, 110], negrilla_ultima=True))
    pat_ant = _patrimonio_anterior(exogena) if exogena else 0
    if pat_ant:
        delta = R(29) - pat_ant
        e.append(Paragraph(
            f"Patrimonio bruto declarado el año anterior: <b>{_fmt(pat_ant)}</b> · variación: "
            f"<b>{'+' if delta >= 0 else ''}{_fmt(delta)}</b> "
            f"({delta / pat_ant * 100:+.1f}%). Una variación no justificada con las rentas "
            f"declaradas puede activar renta por comparación patrimonial (Art. 236 E.T.) — "
            f"documente el origen del incremento.", st["peq"]))

    # ---------------- rentas por cédula ----------------
    e.append(Paragraph("4. Composición de las rentas", st["h2"]))
    filas = [["Cédula / concepto", "Ingresos", "INCRNGO / costos", "Exentas y deduc.", "Renta líquida"]]
    def _fila(nombre_c, ing, incr, exen, liquida):
        filas.append([nombre_c, _fmt(ing), _fmt(incr), _fmt(exen), _fmt(liquida)])
    _fila("Rentas de trabajo", R(32), R(33), R(41), R(42))
    if R(43): _fila("Honorarios", R(43), R(44) + R(45), R(53), R(57))
    if R(58): _fila("Rentas de capital", R(58), R(59) + R(60), R(69), R(73))
    if R(74): _fila("Rentas no laborales", R(74), R(76) + R(77), R(86), R(90))
    if R(99): _fila("Pensiones", R(99), R(100), R(102), R(103))
    div = R(104) + R(107) + R(108) + R(109)
    if div: _fila("Dividendos", div, R(105), R(110), R(106) + R(107) + R(108))
    filas.append(["Renta líquida gravable (cédula general + pensiones)", "", "",
                  "", _fmt(R(97) + R(103))])
    e.append(_tabla(filas, [None, 78, 82, 80, 84], negrilla_ultima=True))
    ded_gmf_txt = ""
    if getattr(datos, "gmf_pagado", 0) > 0:
        ded_gmf_txt = (f" Deducción 4×1000 (GMF, 50% Art. 115, en rentas de trabajo): "
                       f"{_fmt(round(datos.gmf_pagado * 0.5))}.")
    e.append(Paragraph(
        f"Límite de rentas exentas y deducciones aplicado (Art. 336 E.T.): menor entre el 40% de la "
        f"base y 1.340 UVT ({_fmt(p.a_pesos(1340))}). Deducción por dependientes (R139): {_fmt(R(139))}. "
        f"Deducción 1% factura electrónica (R28): {_fmt(R(28))}.{ded_gmf_txt}", st["peq"]))

    # ---------------- terceros principales ----------------
    if exogena:
        top = _top_terceros(exogena, {32, 43, 58, 74, 99, 104})
        if top:
            e.append(Paragraph("4b. Principales ingresos reportados por terceros", st["h2"]))
            filas_t = [["Quién reportó", "Concepto", "Valor"]]
            filas_t += [[a, b, _fmt(v)] for a, b, v in top]
            filas_t.append(["TOTAL ingresos reportados (principales)", "",
                            _fmt(sum(v for *_, v in top))])
            e.append(_tabla(filas_t, [170, None, 90], negrilla_ultima=True))
            e.append(Paragraph(
                "Fuente: información exógena DIAN. Si un tercero reportó mal un valor, se "
                "solicita la corrección al informante antes de declarar.", st["peq"]))

    # ---------------- liquidación ----------------
    e.append(Paragraph("5. Liquidación del impuesto", st["h2"]))
    filas = [["Concepto", "Renglón", "Valor"],
             ["Impuesto sobre rentas líquidas (tabla Art. 241)", "116/117", _fmt(R(116) + R(117))]]
    if R(118) + R(119) + R(120):
        filas.append(["Impuesto sobre dividendos", "118–120", _fmt(R(118) + R(119) + R(120))])
    if R(125):
        filas.append(["(−) Descuentos tributarios", "125", _fmt(R(125))])
    filas.append(["Impuesto neto de renta", "126", _fmt(R(126))])
    if R(127):
        filas.append(["Impuesto de ganancias ocasionales", "127", _fmt(R(127))])
    filas.append(["Total impuesto a cargo", "129", _fmt(R(129))])
    filas += [
        ["(−) Retenciones que le practicaron", "132", _fmt(R(132))],
        ["(−) Saldo a favor y anticipo del año anterior", "130–131", _fmt(R(130) + R(131))],
        ["(+) Anticipo de renta año siguiente", "133", _fmt(R(133))],
    ]
    if R(135):
        filas.append(["(+) Sanciones", "135", _fmt(R(135))])
    e.append(_tabla(filas, [None, 70, 110]))
    e.append(Spacer(1, 8))

    # ---------------- resultado ----------------
    a_pagar = R(136) > 0
    color = ROJO if a_pagar else VERDE
    texto = ("TOTAL SALDO A PAGAR (R136): " + _fmt(R(136))) if a_pagar else \
            ("TOTAL SALDO A FAVOR (R137): " + _fmt(R(137)))
    caja = Table([[Paragraph(f"<b>{texto}</b>",
                             ParagraphStyle("res", fontSize=13, textColor=colors.white,
                                            alignment=TA_CENTER, leading=17))]],
                 colWidths=[None])
    caja.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), color),
                              ("TOPPADDING", (0, 0), (-1, -1), 9),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
                              ("ROUNDEDCORNERS", [6, 6, 6, 6])]))
    e.append(caja)

    # ---------------- retenciones por tercero ----------------
    if exogena:
        rets = _retenciones_por_tercero(exogena)
        if rets:
            e.append(Paragraph("5b. Retenciones certificadas por terceros (R132)", st["h2"]))
            filas_r = [["Agente retenedor", "Retención"]]
            filas_r += [[a, _fmt(v)] for a, v in rets]
            filas_r.append(["Total retenciones tomadas en la declaración", _fmt(R(132))])
            e.append(_tabla(filas_r, [None, 110], negrilla_ultima=True))
            e.append(Paragraph(
                "Solicite el certificado de retención de cada agente para soportar el valor "
                "tomado (la exógena orienta, el certificado soporta).", st["peq"]))

    # ---------------- pronunciamiento profesional ----------------
    # Si el contador corrigió valores frente a lo reportado por terceros
    # (certificados físicos ≠ exógena DIAN), la diferencia queda expresa como
    # su pronunciamiento profesional.
    correcciones = []
    if liq_base is not None:
        etiquetas = {
            29: "Patrimonio bruto", 30: "Deudas",
            32: "Ingresos rentas de trabajo", 33: "INCRNGO rentas de trabajo",
            38: "Intereses de vivienda", 43: "Ingresos honorarios",
            58: "Ingresos rentas de capital", 59: "INCRNGO rentas de capital",
            74: "Ingresos rentas no laborales", 99: "Ingresos por pensiones",
            132: "Retenciones en la fuente",
        }
        for reng, nombre_r in etiquetas.items():
            antes, ahora = liq_base.r(reng), R(reng)
            if abs(antes - ahora) > 1000:
                correcciones.append((reng, nombre_r, antes, ahora))
    if correcciones or (observaciones or "").strip():
        e.append(Paragraph("Pronunciamiento profesional del contador", st["h2"]))
    if correcciones:
        e.append(Paragraph(
            "Cotejados los certificados y soportes físicos del cliente contra la "
            "información exógena reportada por terceros a la DIAN, el contador "
            "corrigió los siguientes valores (prevalecen los soportes):", st["normal"]))
        filas_c = [["Renglón", "Concepto", "Según exógena DIAN", "Valor del contador", "Diferencia"]]
        for reng, nombre_r, antes, ahora in correcciones:
            filas_c.append([str(reng), nombre_r, _fmt(antes), _fmt(ahora), _fmt(ahora - antes)])
        e.append(_tabla(filas_c, [45, None, 88, 88, 80]))
        e.append(Paragraph(
            "Nota: la información exógena es referencial y no exime de declarar la "
            "realidad económica; las diferencias se soportan en los certificados "
            "aportados por el cliente.", st["peq"]))
    if (observaciones or "").strip():
        for linea in observaciones.strip().splitlines():
            if linea.strip():
                e.append(Paragraph(f"• {linea.strip()}", st["normal"]))

    # ---------------- advertencias y notas ----------------
    if liq.advertencias or (exogena and exogena.advertencias):
        e.append(Paragraph("6. Puntos de atención", st["h2"]))
        for a in (liq.advertencias + (exogena.advertencias if exogena else []))[:10]:
            e.append(Paragraph(f"⚠ {a}", st["alerta"]))

    doc.build(e)
    return ruta


def generar_resumen_excel(
    ruta: Path,
    datos: DatosDeclaracion,
    liq: Liquidacion,
    p: Parametros,
    exogena: Optional[ResultadoExogena] = None,
    razones_obligado=None,
    preparado_por: str = "",
    fecha_lim=None,
) -> Path:
    """Papel de trabajo en Excel (misma info del resumen ejecutivo), por cliente.

    Reutiliza los mismos helpers que el PDF (_partidas_renglon, _top_terceros,
    _retenciones_por_tercero) para que los números coincidan exactos."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    R = liq.r
    con = datos.contribuyente
    nombre = " ".join(x for x in (con.primer_nombre, con.otros_nombres,
                                  con.primer_apellido, con.segundo_apellido) if x) \
        or (exogena.nombre if exogena else "")
    NEG, HDR = "1E2432", PatternFill("solid", fgColor="1E2432")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Papel de trabajo 210"
    fila_actual = [1]

    def _cel(txt, col=1, bold=False, size=11, color="000000", fill=None, white=False):
        c = ws.cell(row=fila_actual[0], column=col, value=txt)
        c.font = Font(name="Arial", size=size, bold=bold,
                      color=("FFFFFF" if white else color))
        if fill:
            c.fill = fill
        if isinstance(txt, (int, float)):
            c.number_format = "#,##0"
        return c

    def salto(n=1):
        fila_actual[0] += n

    def titulo(t):
        _cel(t, bold=True, size=12, color=NEG); salto()

    def encab(cols):
        for i, t in enumerate(cols, 1):
            _cel(t, col=i, bold=True, white=True, fill=HDR)
        salto()

    def fila(cols, bold=False):
        for i, t in enumerate(cols, 1):
            _cel(t, col=i, bold=bold)
        salto()

    _cel("Resumen Ejecutivo — Declaración de Renta (Formulario 210)",
         bold=True, size=14, color=NEG); salto()
    _cel(f"Año Gravable {p.anio_gravable} · UVT {int(p.uvt)}"
         + (f" · Preparado por {preparado_por}" if preparado_por else ""),
         size=10, color="808080"); salto(2)

    titulo("1. Datos del declarante")
    fila(["Contribuyente", nombre])
    fila(["Identificación", con.nit or ""])
    if getattr(con, "dv", ""):
        fila(["DV", con.dv])
    salto()

    titulo("2. Patrimonio")
    encab(["Bien / derecho (reportado por)", "Detalle", "Valor"])
    bienes = _partidas_renglon(exogena, 29) if exogena else []
    for a, b, v in bienes:
        fila([a, b, round(v)])
    suma_rep = sum(v for *_, v in bienes)
    if R(29) - suma_rep > 0.5:
        fila(["Otros activos no reportados en exógena",
              "(vehículos, inmuebles, efectivo, ajustes)", round(R(29) - suma_rep)])
    fila(["TOTAL PATRIMONIO BRUTO (R29)", "", round(R(29))], bold=True)
    salto()
    encab(["Deuda (reportada por)", "Detalle", "Valor"])
    deudas = _partidas_renglon(exogena, 30) if exogena else []
    for a, b, v in deudas:
        fila([a, b, round(v)])
    suma_d = sum(v for *_, v in deudas)
    if R(30) - suma_d > 0.5:
        fila(["Otras deudas no reportadas en exógena", "", round(R(30) - suma_d)])
    fila(["TOTAL DEUDAS (R30)", "", round(R(30))], bold=True)
    salto()
    fila(["Patrimonio bruto (R29)", "", round(R(29))])
    fila(["(−) Deudas (R30)", "", round(R(30))])
    fila(["Patrimonio líquido (R31)", "", round(R(31))], bold=True)
    salto()

    titulo("3. Composición de las rentas")
    encab(["Cédula / concepto", "Ingresos", "INCRNGO/costos", "Exentas y deduc.", "Renta líquida"])

    def ced(nom, ing, incr, exen, liqd):
        fila([nom, round(ing), round(incr), round(exen), round(liqd)])

    ced("Rentas de trabajo", R(32), R(33), R(41), R(42))
    if R(43):
        ced("Honorarios", R(43), R(44) + R(45), R(53), R(57))
    if R(58):
        ced("Rentas de capital", R(58), R(59) + R(60), R(69), R(73))
    if R(74):
        ced("Rentas no laborales", R(74), R(76) + R(77), R(86), R(90))
    if R(99):
        ced("Pensiones", R(99), R(100), R(102), R(103))
    fila(["Renta líquida gravable (general + pensiones)", "", "", "", round(R(97) + R(103))], bold=True)
    fila(["Deducción dependientes (R139)", "", "", "", round(R(139))])
    fila(["Deducción 1% factura electrónica (R28)", "", "", "", round(R(28))])
    salto()

    if exogena:
        top = _top_terceros(exogena, {32, 43, 58, 74, 99, 104})
        if top:
            titulo("3b. Principales ingresos reportados por terceros")
            encab(["Quién reportó", "Concepto", "Valor"])
            for a, b, v in top:
                fila([a, b, round(v)])
            fila(["TOTAL ingresos reportados (principales)", "",
                  round(sum(v for *_, v in top))], bold=True)
            salto()

    titulo("4. Liquidación del impuesto")
    encab(["Concepto", "Renglón", "Valor"])
    fila(["Impuesto sobre rentas líquidas (Art. 241)", "116/117", round(R(116) + R(117))])
    fila(["Impuesto neto de renta", "126", round(R(126))])
    fila(["Total impuesto a cargo", "129", round(R(129))], bold=True)
    fila(["(−) Retenciones que le practicaron", "132", round(R(132))])
    fila(["(−) Saldo a favor y anticipo del año anterior", "130-131", round(R(130) + R(131))])
    fila(["(+) Anticipo de renta año siguiente", "133", round(R(133))])
    if R(136) > 0:
        fila(["TOTAL SALDO A PAGAR (R136)", "", round(R(136))], bold=True)
    else:
        fila(["TOTAL SALDO A FAVOR (R137)", "", round(R(137))], bold=True)
    salto()

    if exogena:
        rets = _retenciones_por_tercero(exogena)
        if rets:
            titulo("4b. Retenciones certificadas por terceros (R132)")
            encab(["Agente retenedor", "Retención"])
            for a, v in rets:
                fila([a, round(v)])
            fila(["Total retenciones tomadas en la declaración", round(R(132))], bold=True)

    for coln, w in (("A", 46), ("B", 30), ("C", 16), ("D", 16), ("E", 16)):
        ws.column_dimensions[coln].width = w
    wb.save(str(ruta))
    return ruta


def rellenar_papel_trabajo(entrada, datos, liq, p, exogena=None) -> bytes:
    """Recibe el papel de trabajo del contador (Excel con años anteriores) y le
    AGREGA la columna del año actual, conservando el historial y recalculando los
    totales. Empareja cada entidad por nombre; las nuevas las agrega como filas."""
    import io as _io
    import re
    import unicodedata
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    R = liq.r
    data = entrada.read() if hasattr(entrada, "read") else (
        Path(entrada).read_bytes() if not isinstance(entrada, bytes) else entrada)

    def _norm(s):
        s = str(s or "").upper()
        s = "".join(c for c in unicodedata.normalize("NFD", s)
                    if unicodedata.category(c) != "Mn")
        s = re.sub(r"[^A-Z0-9]", "", s)                       # puntuación/espacios PRIMERO
        s = re.sub(r"(SAS|SA|SCA|SENC|LTDA|EU)$", "", s)      # sufijo legal ('S.A.'->'SA'->'')
        return s

    def _agg(parts):
        d = {}
        for a, _b, v in parts:
            d[a] = d.get(a, 0) + v
        return list(d.items())

    def _anio_col(v):
        """Año de un encabezado de columna: numérico (2024) o texto ('AÑO 2024')."""
        if isinstance(v, (int, float)) and 1990 <= v <= 2100:
            return int(v)
        m = re.search(r"(19|20)\d{2}", str(v or ""))
        return int(m.group(0)) if m else None

    patr = _agg(_partidas_renglon(exogena, 29)) if exogena else []
    deud = _agg(_partidas_renglon(exogena, 30)) if exogena else []
    ing = _agg([(q, "", v) for q, _c, v in _top_terceros(exogena, {32, 43, 58, 74, 99, 104})]) \
        if exogena else []
    val25 = max(0.0, R(41) - R(28) - R(139))
    costos = [("SEGURIDADSOCIAL", R(33)), ("INCRNGO", R(33)), ("BENEFICIODIAN", R(28)),
              ("1FE", R(28)), ("FACTURAELECTR", R(28)), ("25", val25)]
    ret = _retenciones_por_tercero(exogena) if exogena else []

    wb = openpyxl.load_workbook(_io.BytesIO(data))
    wbv = openpyxl.load_workbook(_io.BytesIO(data), data_only=True)
    ws, wsv = wb.worksheets[0], wbv.worksheets[0]
    # aplanar fórmulas -> su valor cacheado (para que insertar filas no rompa nada)
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                c.value = wsv[c.coordinate].value

    NAVY = PatternFill("solid", fgColor="1E2432")
    BAND = PatternFill("solid", fgColor="F4F1EA")
    thin = Side(style="thin", color="D9CFB8")
    BOR = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fill(titulo, items, kw=False):
        titset = ((titulo,) if isinstance(titulo, str) else tuple(titulo))
        titset = tuple(t.upper() for t in titset)
        ft = None
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip().upper() in titset:
                ft = r
                break
        if not ft:
            return
        fh = ft + 1
        # columnas que YA son de un año (numérico o texto 'AÑO 2024') — historial
        ycols = [(c, _anio_col(ws.cell(fh, c).value))
                 for c in range(3, ws.max_column + 2)]
        ycols = [(c, y) for c, y in ycols if y is not None]
        anio = int(getattr(p, "anio_gravable", 2025))      # año gravable REAL (2025)
        # si ya existe la columna de ESE año se reusa; si no, se AGREGA una nueva a
        # la derecha SIN tocar los años anteriores (es un histórico).
        col = next((c for c, y in ycols if y == anio), None)
        if col is None:
            col = (max(c for c, _ in ycols) + 1) if ycols else 3
        ftot = None
        for r in range(fh + 1, ws.max_row + 2):
            if str(ws.cell(r, 1).value or "").upper().startswith("TOTAL"):
                ftot = r
                break
        if not ftot:
            return
        ws.cell(fh, col, anio)                             # SIEMPRE el año gravable, no max+1
        if kw:
            for r in range(fh + 1, ftot):
                clave = _norm(ws.cell(r, 1).value) + _norm(ws.cell(r, 2).value)
                for k, v in items:
                    if k in clave:
                        ws.cell(r, col, round(v))
                        break
        else:
            idx = {}
            for r in range(fh + 1, ftot):
                if ws.cell(r, 1).value:
                    idx.setdefault(_norm(ws.cell(r, 1).value), r)
            ult = max(range(fh + 1, ftot)) if ftot > fh + 1 else fh
            for nombre, valor in items:
                if not valor:
                    continue
                nn = _norm(nombre)
                r = idx.get(nn)
                if r is None:
                    # el contador suele abreviar ('RAPPIPAY' vs 'RAPPIPAY COMPAÑÍA…'):
                    # empareja si su nombre corto (≥6) es prefijo del reportado.
                    for k, rr in idx.items():
                        if len(k) >= 6 and (nn.startswith(k) or k.startswith(nn)):
                            r = rr
                            break
                if r is None:
                    r = ult + 1
                    if r >= ftot:
                        ws.insert_rows(ftot)
                        ftot += 1
                    ws.cell(r, 1, nombre)
                    ult = r
                    idx[_norm(nombre)] = r
                ws.cell(r, col, round(valor))
        # rellenar con 0 las celdas vacías del año (todos los campos llenos)
        for r in range(fh + 1, ftot):
            if ws.cell(r, 1).value and ws.cell(r, col).value in (None, ""):
                ws.cell(r, col, 0)
        L = get_column_letter(col)
        ws.cell(ftot, col, f"=SUM({L}{fh + 1}:{L}{ftot - 1})")
        # tabla elegante en la columna nueva
        hc = ws.cell(fh, col)
        hc.font = Font(name="Arial", bold=True, color="FFFFFF")
        hc.fill = NAVY
        for i, r in enumerate(range(fh + 1, ftot)):
            cc = ws.cell(r, col)
            cc.number_format = "#,##0"
            cc.border = BOR
            if i % 2:
                cc.fill = BAND
        tc = ws.cell(ftot, col)
        tc.font = Font(name="Arial", bold=True)
        tc.number_format = "#,##0"
        tc.border = BOR

    _fill("PATRIMONIO", patr)
    _fill(("PASIVOS", "DEUDAS"), deud)
    _fill("INGRESOS", ing)
    _fill("COSTOS Y DEDUCCIONES", costos, kw=True)
    _fill("RETENCIONES", ret)

    # ---- Formato de marca Tributando en TODA la hoja ----
    from openpyxl.styles import Alignment
    TITULOS = ("PATRIMONIO", "DEUDAS", "PASIVOS", "INGRESOS", "COSTOS Y DEDUCCIONES",
               "COSTOS", "DEDUCCIONES", "LIQUIDACION", "LIQUIDACIÓN", "RENTA",
               "GANANCIAS", "DONACIONES", "RETENCIONES", "DECLARACION", "DECLARACIÓN")
    HEADERS = ("DESCRIPCION", "DESCRIPCIÓN", "NOMBRE", "IDENTIFICACION",
               "IDENTIFICACIÓN", "CONCEPTO", "IDENTIFICACIÓN DEL BIEN", "RENGLON", "RENGLÓN")
    # última columna CON datos (para no bordear columnas vacías a la derecha)
    maxc = 3
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value not in (None, ""):
                maxc = max(maxc, c)
    izq = Alignment(horizontal="left"); der = Alignment(horizontal="right")
    cen = Alignment(horizontal="center")
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "").strip().upper()
        b = str(ws.cell(r, 2).value or "").strip().upper()
        fila_vacia = not any(ws.cell(r, c).value not in (None, "") for c in range(1, maxc + 1))
        es_titulo = (not any(ch.isdigit() for ch in a)) and any(
            a == t or (a.startswith(t) and len(a) - len(t) < 12) for t in TITULOS)
        es_header = a in HEADERS or b in HEADERS
        es_total = a.startswith("TOTAL")
        if fila_vacia and not es_titulo:
            continue
        for c in range(1, maxc + 1):
            cell = ws.cell(r, c)
            cell.border = BOR
            if es_titulo:
                cell.fill = NAVY
                cell.font = Font(name="Arial", bold=True, color="C9A75A", size=11)
            elif es_header:
                cell.fill = NAVY
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.alignment = cen if c >= 3 else izq
            elif es_total:
                cell.fill = BAND
                cell.font = Font(name="Arial", bold=True)
            else:
                if r % 2 == 0:
                    cell.fill = BAND
                cell.font = Font(name="Arial", size=10)
            if isinstance(cell.value, (int, float)):
                # los años (encabezado) van sin separador de miles: 2025, no 2.025
                if es_header or (1990 <= cell.value <= 2100 and float(cell.value).is_integer()):
                    cell.number_format = "0"
                else:
                    cell.number_format = "#,##0"
                    cell.alignment = der
    for c in range(3, maxc + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
