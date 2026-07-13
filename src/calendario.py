"""Calendario tributario: fecha límite de declaración según los últimos
dos dígitos del NIT/cédula (hoja 'Calendario Tributario' de la plantilla).
"""
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional

import openpyxl

_CACHE: Optional[Dict[str, date]] = None


def cargar_calendario(plantilla: Path) -> Dict[str, date]:
    """Lee el mapa {últimos 2 dígitos → fecha límite} de la plantilla (B21:C120)."""
    global _CACHE
    if _CACHE is not None:
        return _CACHE
    calendario: Dict[str, date] = {}
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(plantilla, data_only=True, read_only=True)
    if "Calendario Tributario" in wb.sheetnames:
        ws = wb["Calendario Tributario"]
        for fila in ws.iter_rows(min_row=21, max_row=120, min_col=2, max_col=3):
            clave, fecha = fila[0].value, fila[1].value
            if clave is None or fecha is None:
                continue
            clave = str(clave).strip().zfill(2)[-2:]
            if isinstance(fecha, datetime):
                calendario[clave] = fecha.date()
    wb.close()
    # Corrige años atípicos por tipeo en la plantilla: los vencimientos de un
    # mismo año gravable caen todos en el mismo año calendario, así que forzamos
    # los outliers al año mayoritario (p. ej. cédulas 63/64 venían con 2025 en vez
    # de 2026). Evita mostrarle a esas personas una fecha límite ya vencida.
    if calendario:
        from collections import Counter
        anio_ok = Counter(f.year for f in calendario.values()).most_common(1)[0][0]
        for clave, f in calendario.items():
            if f.year != anio_ok:
                calendario[clave] = f.replace(year=anio_ok)
    _CACHE = calendario
    return calendario


def fecha_limite(nit: str, plantilla: Path) -> Optional[date]:
    """Fecha límite de declaración y pago para un NIT/cédula."""
    digitos = "".join(c for c in str(nit) if c.isdigit())
    if len(digitos) < 2:
        return None
    return cargar_calendario(plantilla).get(digitos[-2:])
