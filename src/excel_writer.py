"""Escritura del resultado sobre la plantilla Excel (hoja 'FORMULARIO 210').

Copia la plantilla ITGS y escribe:
 - los datos del contribuyente en la hoja 'Datos del contribuyente',
 - el número de dependientes en la hoja 'Dependientes ',
 - los valores calculados directamente en las celdas de cada renglón de la
   hoja 'FORMULARIO 210' (reemplaza las fórmulas por valores para que el
   archivo sea autocontenido y refleje la liquidación del motor),
 - una hoja nueva 'Trazabilidad' con el log de qué filas de la exógena
   alimentaron cada renglón.

El formato de la plantilla (estilos, merges) se preserva porque solo se
tocan valores de celdas existentes.
"""
import shutil
import warnings
from pathlib import Path
from typing import Optional

import openpyxl

from .hojas_detalle import llenar_hojas_detalle
from .modelos import DatosDeclaracion, Liquidacion, ResultadoExogena
from .parametros import Parametros

HOJA_FORMULARIO = "FORMULARIO 210"
HOJA_DATOS = "Datos del contribuyente"
HOJA_DEPENDIENTES = "Dependientes "

# Mapa renglón → celda en la hoja 'FORMULARIO 210' de la plantilla ITGS.
# Levantado inspeccionando la plantilla: cada renglón tiene su celda de valor.
CELDAS_RENGLON = {
    28: "AB12",
    29: "L13", 30: "S13", 31: "Z13",
    32: "H15", 43: "M15", 58: "S15", 74: "Z15",
    75: "Z16",
    33: "H17", 44: "M17", 59: "S17", 76: "Z17",
    45: "M18", 60: "S18", 77: "Z18",
    34: "H19", 46: "M19", 61: "S19", 78: "Z19",
    62: "S20", 79: "Z20",
    35: "H21", 47: "M21", 63: "S21", 80: "Z21",
    36: "H22", 48: "M22", 64: "S22", 81: "Z22",
    37: "H23", 49: "M23", 65: "S23", 82: "Z23",
    38: "H24", 50: "M24", 66: "S24", 83: "Z24",
    39: "H25", 51: "M25", 67: "S25", 84: "Z25",
    40: "H26", 52: "M26", 68: "S26", 85: "Z26",
    41: "H27", 53: "M27", 69: "S27", 86: "Z27",
    54: "M28", 70: "S28", 87: "Z28",
    55: "M29", 71: "S29", 88: "Z29",
    56: "M30", 72: "S30", 89: "Z30",
    42: "H31", 57: "M31", 73: "S31", 90: "Z31",
    91: "F32", 92: "L32", 93: "S32", 94: "AA32",
    95: "F33", 96: "L33", 97: "S33", 98: "AA33",
    99: "J34", 100: "J35", 101: "J36", 102: "J37", 103: "J38",
    104: "J39", 105: "J40", 106: "J41", 107: "J42", 108: "J43",
    109: "J44", 110: "J45",
    111: "J46",
    112: "J47", 113: "J48", 114: "J49", 115: "J50",
    116: "Y34", 117: "Y35", 118: "Y36", 119: "Y37", 120: "Y38", 121: "Y39",
    122: "U40", 123: "AB40", 124: "U41", 125: "AB41",
    126: "Y42", 127: "Y43", 128: "Y44", 129: "Y45",
    130: "Y46", 131: "Y47", 132: "Y48", 133: "Y49",
    134: "F51", 135: "M51", 136: "U51", 137: "AB51",
    138: "F52", 139: "M52",
    141: "AB52",
}


def escribir_formulario(
    plantilla: Path,
    salida: Path,
    datos: DatosDeclaracion,
    liq: Liquidacion,
    exogena: Optional[ResultadoExogena] = None,
    parametros: Optional[Parametros] = None,
) -> Path:
    """Genera el Excel de salida. Devuelve la ruta escrita."""
    plantilla, salida = Path(plantilla), Path(salida)
    if not plantilla.exists():
        raise FileNotFoundError(f"No existe la plantilla: {plantilla}")
    salida.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(plantilla, salida)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(salida)

    # ---- datos del contribuyente ------------------------------------
    con = datos.contribuyente
    if HOJA_DATOS in wb.sheetnames:
        ws = wb[HOJA_DATOS]
        ws["C6"] = con.nit
        ws["C7"] = con.dv
        ws["C8"] = con.primer_apellido
        ws["C9"] = con.segundo_apellido
        ws["C10"] = con.primer_nombre
        ws["C11"] = con.otros_nombres
        ws["C12"] = con.actividad_economica
        ws["C13"] = "x" if con.es_correccion else ""
        ws["C14"] = con.formulario_anterior

    # ---- dependientes -------------------------------------------------
    if HOJA_DEPENDIENTES in wb.sheetnames:
        ws = wb[HOJA_DEPENDIENTES]
        ws["C3"] = datos.dependientes
        # nombres en B7:B10 (la plantilla deduce por nombre presente); si solo
        # hay conteo se escriben marcadores para no dejar los de ejemplo
        nombres = [n for n in datos.dependientes_detalle if str(n).strip()][:4]
        if not nombres and datos.dependientes > 0:
            nombres = [f"Dependiente {i+1}" for i in range(min(datos.dependientes, 4))]
        for i in range(4):
            # asignación directa: cell(value=None) no borra en openpyxl
            ws.cell(row=7 + i, column=2).value = nombres[i] if i < len(nombres) else None

    # ---- hojas de detalle: datos reales, sin ejemplos de la plantilla --
    llenar_hojas_detalle(wb, datos, liq, exogena, parametros)

    # ---- formulario 210: valores por renglón --------------------------
    if HOJA_FORMULARIO not in wb.sheetnames:
        raise ValueError(f"La plantilla no tiene la hoja '{HOJA_FORMULARIO}'.")
    ws = wb[HOJA_FORMULARIO]
    for renglon, celda in CELDAS_RENGLON.items():
        if renglon in liq.renglones:
            ws[celda] = liq.renglones[renglon]

    # ---- trazabilidad --------------------------------------------------
    if exogena is not None:
        nombre = "Trazabilidad"
        if nombre in wb.sheetnames:
            del wb[nombre]
        tz = wb.create_sheet(nombre)
        tz.append(["Fila exógena", "Renglón asignado", "Detalle", "Valor",
                   "NIT informante", "Informante", "Excluida", "Nota"])
        for p in exogena.partidas:
            tz.append([
                p.fila,
                f"R{p.renglon_asignado}" if p.renglon_asignado else "",
                p.detalle, p.valor, p.informante_nit, p.informante_nombre,
                "sí" if p.excluida else "", p.nota,
            ])
        tz.append([])
        tz.append(["Advertencias de la liquidación:"])
        for a in liq.advertencias:
            tz.append(["", a])
        for col, ancho in zip("ABCDEFGH", (12, 14, 60, 16, 14, 40, 9, 60)):
            tz.column_dimensions[col].width = ancho

    # ---- limpiar valores de EJEMPLO de la plantilla que confunden al contador --
    # La hoja 'G.OCAS' (ganancias ocasionales) trae ejemplos hardcodeados (tipo de
    # bien, años y fechas de compra/venta ficticias). Se borran para que la hoja
    # quede en blanco y no parezcan datos reales del contribuyente.
    if "G.OCAS" in wb.sheetnames:
        wsg = wb["G.OCAS"]
        for celda in ("F11", "F12", "F25", "F26", "F33", "F34", "F39", "F40",
                      "F5", "F6", "F19", "F20"):
            wsg[celda] = None
    if "REAJUST FISCAL" in wb.sheetnames:
        wsr = wb["REAJUST FISCAL"]
        for celda in ("J12", "J70", "J140"):   # fechas de ejemplo hardcodeadas
            wsr[celda] = None

    # ---- Comerciante (PN): CMV y ERI cedulado (base para el formato 2517) ----
    if (datos.inventario_inicial or datos.inventario_final or datos.compras_mercancia
            or datos.activo_vehiculos or datos.activo_maquinaria
            or datos.activo_muebles or datos.activo_equipo_computo
            or datos.depreciacion_manual or datos.activos_fijos):
        _hoja_comerciante(wb, datos, liq)

    wb.save(salida)
    return salida


def escribir_borrador_comerciante(salida: Path, datos: DatosDeclaracion,
                                  liq: Liquidacion) -> Path:
    """Excel ENFOCADO del borrador del comerciante: solo la hoja 'Borrador 210'
    (renglones clave) y la hoja 'Comerciante (CMV-ERI)' con la guía del formato
    2517 — sin los demás papeles de trabajo de la plantilla."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    salida = Path(salida)
    AZUL, ORO_CLARO, GRIS = "1E2432", "E0C584", "F6F4EE"
    thin = Side(style="thin", color="E2DCCC")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Borrador 210"

    # banner de marca (azul marino + dorado)
    ws.append(["Tributando.co  ·  Borrador de renta — Comerciante (PN)"])
    ws.merge_cells("A1:C1")
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A1"].font = Font(bold=True, size=17, color=ORO_CLARO)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    con = datos.contribuyente
    nom = " ".join(x for x in (con.primer_nombre, con.primer_apellido,
                               con.segundo_apellido) if x).strip()
    ws.append([f"{nom or ''}  ·  NIT {con.nit or ''}  ·  Año gravable 2025"])
    ws.merge_cells("A2:C2")
    ws["A2"].font = Font(size=11, color="5A6B7F")
    ws.append([])
    ws.append(["Renglón", "Concepto", "Valor"])
    for c in ws[ws.max_row]:
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.font = Font(color=ORO_CLARO, bold=True, size=12)
        c.border = box
    NOM = {29: "Patrimonio bruto", 30: "Deudas", 31: "Patrimonio líquido",
           32: "Ingresos brutos rentas de trabajo/honorarios", 33: "INCRNGO trabajo",
           36: "Rentas exentas trabajo (incluye 25% Art. 206-10)", 42: "Renta líquida trabajo",
           74: "Ingresos no laborales", 75: "Devoluciones/rebajas", 76: "INCRNGO",
           77: "Costos y deducciones (CMV + depreciación)", 78: "Renta líquida no laboral",
           91: "Renta líquida cédula general", 97: "Renta líquida gravable",
           116: "Impuesto rentas líquidas", 126: "Impuesto neto de renta",
           129: "Total impuesto a cargo",
           130: "Anticipo renta liquidado año gravable anterior",
           131: "Saldo a favor año gravable anterior",
           132: "Retenciones", 133: "Anticipo año gravable siguiente",
           136: "Saldo a pagar", 137: "Saldo a favor"}
    destacar = {97, 129, 136, 137}
    filas_reng = [29, 30, 31]
    if (datos.trabajo.ingresos_brutos or 0) > 0:
        filas_reng += [32, 33, 36, 42]
    filas_reng += [74, 75, 76, 77, 78, 91, 97, 116, 126, 129]
    if (datos.anticipo_anterior or 0) > 0:
        filas_reng.append(130)
    if (datos.saldo_favor_anterior or 0) > 0:
        filas_reng.append(131)
    filas_reng += [132, 133, 136, 137]
    for r in filas_reng:
        if r in liq.renglones:
            ws.append([f"R{r}", NOM.get(r, ""), round(liq.renglones[r])])
            fila = ws[ws.max_row]
            fila[2].number_format = "#,##0"
            for c in fila:
                c.border = box
            if r in destacar:
                for c in fila:
                    c.fill = PatternFill("solid", fgColor=GRIS)
                    c.font = Font(bold=True)
    for col, w in zip("ABC", (14, 50, 24)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    _hoja_comerciante(wb, datos, liq)
    wb.save(salida)
    return salida


def _hoja_comerciante(wb, datos, liq) -> None:
    """Hoja con el costo de ventas por inventarios y el ERI cedulado de la renta
    no laboral — muestra base para diligenciar el formato 2517 (anexo 210)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    nombre = "Comerciante (CMV-ERI)"
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre)
    AZUL, ORO = "1E2432", "C9A75A"

    def titulo(txt):
        ws.append([txt]); ws[ws.max_row][0].font = Font(bold=True, size=14, color=AZUL)

    def encab(cols):
        ws.append(cols)
        for c in ws[ws.max_row]:
            c.fill = PatternFill("solid", fgColor=AZUL)
            c.font = Font(color="E0C584", bold=True, size=11)
            c.alignment = Alignment(horizontal="center")

    def fila(etq, val, bold=False):
        ws.append([etq, round(val)])
        ws[ws.max_row][1].number_format = "#,##0"
        if bold:
            for c in ws[ws.max_row]: c.font = Font(bold=True)

    nl = datos.no_laboral
    compras = datos.compras_mercancia
    cmv = max(0, compras + datos.inventario_inicial - datos.inventario_final)
    # banner de marca
    ws.append(["Tributando.co  ·  Comerciante — CMV, ERI y Conciliación fiscal 2517"])
    ws.merge_cells("A1:D1")
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A1"].font = Font(bold=True, size=15, color="E0C584")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32
    ws.append(["Base para diligenciar el formato 2517 (conciliación fiscal, anexo 210). "
               "La columna contable la ajusta el contador."])
    ws["A2"].font = Font(size=10.5, color="5A6B7F")
    ws.append([])

    titulo("Costo de la mercancía vendida — CMV (Arts. 62/63)")
    encab(["Concepto", "Valor fiscal"])
    fila("Inventario inicial (1-ene)", datos.inventario_inicial)
    fila("(+) Compras del año", compras)
    fila("(−) Inventario final (31-dic)", datos.inventario_final)
    fila("= CMV (costo de ventas)", cmv, bold=True)
    ws.append([])

    # Depreciación de activos fijos (Art. 137): totales rápidos + lista detallada.
    from .motor_calculo import calcular_depreciacion, _TASA_CATEGORIA
    _CAT_NOMBRE = {"vehiculos": "Vehículos", "maquinaria": "Maquinaria",
                   "muebles": "Muebles y enseres", "computo": "Equipo de cómputo",
                   "construcciones": "Construcciones", "otros": "Otros"}
    deps = [("Vehículos (10%)", datos.activo_vehiculos, 0.10),
            ("Maquinaria y equipo (10%)", datos.activo_maquinaria, 0.10),
            ("Muebles y enseres (10%)", datos.activo_muebles, 0.10),
            ("Equipo de cómputo (20%)", datos.activo_equipo_computo, 0.20)]
    dep_total = calcular_depreciacion(datos)
    if dep_total:
        titulo("Depreciación de activos fijos (Art. 137, línea recta)")
        encab(["Activo / categoría", "Costo fiscal", "Depreciación año", "¿En exógena?"])
        for etq, val, tasa in deps:
            if val:
                ws.append([etq, round(val), round(val * tasa), ""])
                for c in ws[ws.max_row][1:3]: c.number_format = "#,##0"
        for a in getattr(datos, "activos_fijos", []):
            if not a.valor:
                continue
            tasa = _TASA_CATEGORIA.get(a.categoria, 0.10)
            etq = f"{a.descripcion or _CAT_NOMBRE.get(a.categoria, a.categoria)} " \
                  f"({_CAT_NOMBRE.get(a.categoria, a.categoria)}, {tasa:.2%})"
            ws.append([etq, round(a.valor), round(a.valor * tasa),
                       "Sí" if a.en_exogena else "No → suma a R29"])
            for c in ws[ws.max_row][1:3]: c.number_format = "#,##0"
        if datos.depreciacion_manual:
            ws.append(["Otra (manual)", "", round(datos.depreciacion_manual), ""])
            ws[ws.max_row][2].number_format = "#,##0"
        ws.append(["Total depreciación del año → costos R77", "", round(dep_total), ""])
        r = ws[ws.max_row]; r[2].number_format = "#,##0"
        for c in r: c.font = Font(bold=True)
        ws.append([])

    titulo("ERI — Cédula de rentas NO laborales (comercial)")
    encab(["Concepto", "Valor fiscal", "Renglón 210"])
    for etq, r in [("Ingresos brutos", 74), ("(−) Devoluciones/rebajas", 75),
                   ("(−) INCRNGO", 76), ("(−) Costos y deducciones (CMV)", 77),
                   ("= Renta líquida no laboral", 78)]:
        ws.append([etq, round(liq.renglones.get(r, 0)), f"R{r}"])
        ws[ws.max_row][1].number_format = "#,##0"
        if r == 78:
            for c in ws[ws.max_row]: c.font = Font(bold=True)
    ws.append([])

    titulo("ESF — Patrimonio (parcial)")
    encab(["Concepto", "Valor fiscal", "Renglón 210"])
    ws.append(["Inventario final (activo)", round(datos.inventario_final), "R29"])
    ws[ws.max_row][1].number_format = "#,##0"
    ws.append(["Patrimonio bruto total", round(liq.renglones.get(29, 0)), "R29"])
    ws[ws.max_row][1].number_format = "#,##0"
    ws.append([])

    # ---- Guía para diligenciar el formato 2517 (F2517 v8) — filas exactas ----
    titulo("GUÍA F2517 — lleve estos valores a la hoja H3 (ERI) del prevalidador")
    encab(["Destino en el formato 2517 (hoja · fila · concepto)", "Valor fiscal",
           "Valor contable (llene)", "Diferencia"])
    guia = [
        ("H3 ERI · fila 40 · Venta de bienes (Al territorio nacional)",
         liq.renglones.get(74, 0)),
        ("H3 ERI · fila 150 · Costo bienes vendidos (comerciantes): Inventario inicial",
         datos.inventario_inicial),
        ("H3 ERI · fila 151 · Costo bienes vendidos (comerciantes): compras locales",
         compras),
        ("H3 ERI · fila 153 · Costo bienes vendidos (comerciantes): Inventario final",
         datos.inventario_final),
        ("H3 ERI · fila 167 · Depreciación propiedades, planta y equipo (Del costo)",
         dep_total),
    ]
    for etq, val in guia:
        f = ws.max_row + 1
        # col C = contable (la llena el contador); col D = diferencia = contable − fiscal
        ws.append([etq, round(val or 0), None, f"=C{f}-B{f}"])
        for i in (1, 2, 3):
            ws[ws.max_row][i].number_format = "#,##0"
    ws.append(["Los activos de la lista van a la hoja H6 (Activos fijos) por su categoría "
               "(Equipos de Transporte, Maquinaria, Enseres, Equipos informáticos…).", ""])
    ws.append([])
    ws.append(["Nota: el formato 2517 pide además la columna CONTABLE (NIIF) y las "
               "diferencias por partida; esta hoja entrega la base fiscal ya clasificada "
               "en la cédula no laboral. El CMV usa el sistema periódico (inv. inicial + "
               "compras − inv. final)."])
    for col, ancho in zip("ABCD", (60, 20, 16, 18)):
        ws.column_dimensions[col].width = ancho
