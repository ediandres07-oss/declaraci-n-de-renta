"""Pruebas del parser de exógena usando el fixture real de Elizabeth."""
import openpyxl
import pytest

from src.exogena_parser import (ExogenaError, _parse_valor, calcular_topes_propios,
                                evaluar_obligacion_declarar, parsear_exogena)
from src.modelos import PartidaExogena
from src.exogena_parser import _clasificar

from .conftest import EXOGENA, FIXTURES


# ---------------- extracción de renglones (regex tolerante) ----------------

@pytest.mark.parametrize("uso, esperados, asignado", [
    ("R132 Retenciones año gravable a declarar", [132], 132),
    ("R30 Deudas", [30], 30),
    ("r 29 patrimonio", [29], 29),
    ("Renglón 58 ingresos", [58], 58),
    ("renglon 74", [74], 74),
    ("Tope 1: ingresos brutos | R58 Ingresos brutos por rentas de capital", [58], 58),
    ("Tope 1: ingresos brutos | R58 ... | R59 Ingresos no constitutivos", [58, 59], 58),
    ("R33 Ingresos no constitutivos de renta o R100 Ingresos no constitutivos", [33, 100], 33),
    ("Tope 2. Patrimonio| R29 Patrimonio bruto", [29], 29),
])
def test_extraccion_renglones(uso, esperados, asignado):
    p = PartidaExogena(fila=1, informante_nit="", informante_nombre="", informado_nit="",
                       informado_nombre="", detalle="x", valor=1.0, uso_sugerido=uso)
    _clasificar(p)
    assert p.renglones == esperados
    assert p.renglon_asignado == asignado


def test_texto_sin_codigo_aporte_obligatorio_va_a_r33():
    p = PartidaExogena(fila=1, informante_nit="", informante_nombre="", informado_nit="",
                       informado_nombre="", detalle="Aporte obligatorio pensiones",
                       valor=1.0, uso_sugerido="Ingresos no constitutivos de renta")
    _clasificar(p)
    assert p.renglon_asignado == 33


def test_topes_detectados_en_uso():
    p = PartidaExogena(fila=1, informante_nit="", informante_nombre="", informado_nit="",
                       informado_nombre="", detalle="x", valor=1.0,
                       uso_sugerido="TOPE 4. Consignaciones e inversiones")
    _clasificar(p)
    assert p.topes == [4]
    assert p.renglon_asignado is None


# ---------------- parseo de montos con formatos raros ----------------------

@pytest.mark.parametrize("crudo, esperado", [
    (1234567, 1234567.0),
    ("1.234.567", 1234567.0),
    ("1,234,567.89", 1234567.89),
    ("1.234.567,89", 1234567.89),
    ("$ 500.000", 500000.0),
    ("", None),
    ("N/A", None),
])
def test_parse_valor(crudo, esperado):
    assert _parse_valor(crudo) == esperado


# ---------------- fixture real ---------------------------------------------

def test_metadatos_elizabeth(exogena_elizabeth):
    r = exogena_elizabeth
    assert r.anio == 2025
    assert r.identificacion == "44004730"
    assert "ELIZABETH" in r.nombre
    assert r.tipo_documento.startswith("C")


def test_totales_por_renglon_elizabeth(exogena_elizabeth):
    tot = exogena_elizabeth.total_por_renglon()
    # verificados a mano contra el reporte
    assert tot[32] == 86_884_216      # salarios + otros pagos + prestaciones + cesantías pagadas + fondo
    assert tot[33] == 3_992_610       # aportes obligatorios salud (1.897.503) + pensión (2.095.107)
    assert tot[30] == 658_958
    assert tot[132] == 1_362_514
    assert tot[58] == 38_242_290
    assert tot[74] == 644_700
    assert tot[28] == 8_128_113
    assert tot[29] == 458_522_501


def test_topes_dian_leidos(exogena_elizabeth):
    t = exogena_elizabeth.topes_dian
    assert t["ingresos"] == 125_771_208
    assert t["patrimonio"] == 515_035_008
    assert t["consumos_tc"] == 15_666_577
    assert t["consignaciones"] == 1_297_938_176
    assert t["compras"] == 8_128_113


def test_topes_propios_coinciden_con_dian(exogena_elizabeth):
    """La reagregación propia valida el resumen de la DIAN (tolerancia por redondeos)."""
    propios = calcular_topes_propios(exogena_elizabeth)
    dian = exogena_elizabeth.topes_dian
    for clave, valor_dian in dian.items():
        assert propios[clave] == pytest.approx(valor_dian, abs=100), clave


def test_obligacion_declarar_elizabeth(exogena_elizabeth, parametros):
    razones = evaluar_obligacion_declarar(exogena_elizabeth.topes_dian, parametros)
    texto = " ".join(razones)
    assert "Patrimonio" in texto
    assert "Ingresos" in texto
    assert "Consignaciones" in texto
    assert len(razones) == 3  # consumos TC y compras no superan el tope


def test_un_solo_nit_sin_advertencia(exogena_elizabeth):
    assert not any("más de un NIT" in a for a in exogena_elizabeth.advertencias)


# ---------------- manejo de errores -----------------------------------------

def test_archivo_inexistente():
    with pytest.raises(ExogenaError, match="no existe"):
        parsear_exogena(FIXTURES / "no_existe.xlsx")


def test_archivo_corrupto(tmp_path):
    malo = tmp_path / "malo.xlsx"
    malo.write_bytes(b"esto no es un xlsx")
    with pytest.raises(ExogenaError, match="No se pudo abrir"):
        parsear_exogena(malo)


def test_sin_columna_uso(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append(["NIT", "Nombre / Razon Social", "NIT", "Nombre", "Detalle", "Valor"])
    ws.append([1, "X", 2, "Y", "algo", 100])
    ruta = tmp_path / "sin_uso.xlsx"
    wb.save(ruta)
    with pytest.raises(ExogenaError, match="Uso declaración Sugerida"):
        parsear_exogena(ruta)


def test_hoja_distinta_y_columnas_desordenadas(tmp_path):
    """Debe soportar otro nombre de hoja, otro orden de columnas y montos como texto."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos2024"
    ws.append(["Reporte año 2024"])
    ws.append(["Identificación:", None, "123"])
    ws.append(["Valor", "Uso declaración Sugerida", "Detalle", "NIT",
               "Nombre / Razon Social", "NIT", "Nombre/Razon Social reportada por el tercero"])
    ws.append(["1.500.000", "R32 Ingresos", "Pagos por salarios", 900, "EMPRESA", "123", "PERSONA"])
    ws.append([])  # fila vacía intercalada
    ws.append(["no-numerico", "R30 Deudas", "Deuda rara", 901, "BANCO", "123", "PERSONA"])
    ruta = tmp_path / "otra.xlsx"
    wb.save(ruta)
    r = parsear_exogena(ruta)
    assert r.total_por_renglon() == {32: 1_500_000}
    assert any("no es numérico" in a for a in r.advertencias)
    assert any("Reporte" in a for a in r.advertencias)  # advierte nombre de hoja distinto


def test_multiples_nits_advertencia(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append(["NIT", "Nombre / Razon Social", "NIT",
               "Nombre/Razon Social reportada por el tercero", "Detalle", "Valor",
               "Uso declaración Sugerida", "Informacion Adicional"])
    ws.append([1, "A", "111", "P1", "pago", 10, "R32 Ingresos", ""])
    ws.append([2, "B", "222", "P2", "pago", 20, "R32 Ingresos", ""])
    ruta = tmp_path / "dos_nits.xlsx"
    wb.save(ruta)
    r = parsear_exogena(ruta)
    assert any("más de un NIT" in a for a in r.advertencias)


# ---------------- beneficiario económico / titularidad ----------------------

def _partida_con_info(info_adicional, detalle="Valor avalúo catastral", valor=100_000_000.0):
    from src.exogena_parser import _ajustar_beneficiario_economico
    p = PartidaExogena(fila=1, informante_nit="", informante_nombre="", informado_nit="",
                       informado_nombre="", detalle=detalle, valor=valor,
                       uso_sugerido="Tope 2. Patrimonio | R29 Patrimonio bruto",
                       info_adicional=info_adicional)
    _clasificar(p)
    aviso = _ajustar_beneficiario_economico(p)
    return p, aviso


def test_participacion_parcial_ajusta_valor():
    p, aviso = _partida_con_info(
        "Porcentaje de Participación: 50.00 | Matricula: 123 | Número Propietarios: 2")
    assert p.participacion == 50.0
    assert p.valor == 50_000_000
    assert p.valor_reportado == 100_000_000
    assert "50" in p.nota and aviso is not None


def test_participacion_total_no_ajusta():
    p, aviso = _partida_con_info(
        "Porcentaje de Participación: 100.00 | Matricula: 123 | Número Propietarios: 1")
    assert p.participacion == 100.0
    assert p.valor == 100_000_000
    assert p.valor_reportado is None
    assert aviso is None


def test_varios_propietarios_sin_porcentaje_solo_advierte():
    p, aviso = _partida_con_info("Número Propietarios: 3")
    assert p.num_propietarios == 3
    assert p.valor == 100_000_000          # no se ajusta sin % informado
    assert "titular principal" in p.nota
    assert aviso is not None and "3 propietarios" in aviso


def test_cotitular_marca_revision():
    p, aviso = _partida_con_info("Tipo de Titular: Cotitular",
                                 detalle="Saldo cuentas bancarias (Titular Secundario)")
    assert "beneficiario económico" in p.nota or "cotitular" in p.nota.lower()
    assert aviso is not None


def test_fixture_elizabeth_participaciones_100(exogena_elizabeth):
    """En el fixture todas las participaciones son 100%/1 propietario: nada cambia."""
    ajustadas = [p for p in exogena_elizabeth.partidas if p.valor_reportado is not None]
    assert ajustadas == []
    con_participacion = [p for p in exogena_elizabeth.partidas if p.participacion is not None]
    assert len(con_participacion) == 3      # los 3 avalúos catastrales del Municipio de Bello
    assert all(p.participacion == 100.0 for p in con_participacion)
