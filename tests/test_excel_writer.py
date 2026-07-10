"""Pruebas de la escritura del Excel de salida sobre la plantilla ITGS."""
import warnings

import openpyxl
import pytest

from src.entrevista import Sesion, mapear_exogena_a_datos
from src.excel_writer import CELDAS_RENGLON, escribir_formulario
from src.motor_calculo import calcular

from .conftest import PLANTILLA


@pytest.fixture()
def salida_elizabeth(tmp_path, exogena_elizabeth, parametros):
    datos = mapear_exogena_a_datos(exogena_elizabeth)
    datos.contribuyente.dv = "7"
    datos.contribuyente.primer_apellido = "GIRALDO"
    datos.contribuyente.primer_nombre = "ELIZABETH"
    liq = calcular(datos, parametros)
    ruta = escribir_formulario(PLANTILLA, tmp_path / "salida.xlsx", datos, liq,
                               exogena_elizabeth)
    return ruta, liq, datos


def _abrir(ruta):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(ruta)


def test_renglones_escritos_en_formulario(salida_elizabeth):
    ruta, liq, _ = salida_elizabeth
    ws = _abrir(ruta)["FORMULARIO 210"]
    for renglon, celda in CELDAS_RENGLON.items():
        if renglon in liq.renglones:
            assert ws[celda].value == liq.renglones[renglon], f"R{renglon} en {celda}"


def test_datos_contribuyente_escritos(salida_elizabeth):
    ruta, _, datos = salida_elizabeth
    ws = _abrir(ruta)["Datos del contribuyente"]
    assert ws["C6"].value == datos.contribuyente.nit
    assert ws["C8"].value == "GIRALDO"
    assert ws["C10"].value == "ELIZABETH"


def test_hoja_trazabilidad(salida_elizabeth, exogena_elizabeth):
    ruta, _, _ = salida_elizabeth
    wb = _abrir(ruta)
    assert "Trazabilidad" in wb.sheetnames
    tz = wb["Trazabilidad"]
    # encabezado + una fila por partida como mínimo
    assert tz.max_row >= len(exogena_elizabeth.partidas) + 1
    assert tz["A1"].value == "Fila exógena"


def test_formato_preservado(salida_elizabeth):
    """Las demás hojas de la plantilla siguen presentes."""
    ruta, _, _ = salida_elizabeth
    wb = _abrir(ruta)
    for hoja in ("FORMULARIO 210", "Pat bruto", "Deudas", "R.trabajo y honorarios",
                 "impto renta", "anticipo", "retefuente"):
        assert hoja in wb.sheetnames


def test_plantilla_inexistente(tmp_path, parametros, exogena_elizabeth):
    datos = mapear_exogena_a_datos(exogena_elizabeth)
    liq = calcular(datos, parametros)
    with pytest.raises(FileNotFoundError):
        escribir_formulario(tmp_path / "nope.xlsx", tmp_path / "out.xlsx", datos, liq)


# ------------------- sesión (persistir/retomar entrevista) ------------------

def test_sesion_roundtrip(tmp_path, exogena_elizabeth):
    ses = Sesion(exogena=exogena_elizabeth,
                 datos=mapear_exogena_a_datos(exogena_elizabeth),
                 seccion_actual=3, resumen_confirmado=True)
    ruta = tmp_path / "sesion.json"
    ses.guardar(ruta)
    otra = Sesion.cargar(ruta)
    assert otra.seccion_actual == 3
    assert otra.resumen_confirmado is True
    assert otra.datos.trabajo.ingresos_brutos == ses.datos.trabajo.ingresos_brutos
    assert len(otra.exogena.partidas) == len(exogena_elizabeth.partidas)
    assert otra.exogena.total_por_renglon() == exogena_elizabeth.total_por_renglon()


def test_nombres_dependientes_en_hoja(tmp_path, exogena_elizabeth, parametros):
    from src.motor_calculo import calcular as _calc
    datos = mapear_exogena_a_datos(exogena_elizabeth)
    datos.dependientes_detalle = ["Ana María", "Luis"]
    datos.dependientes = 2
    liq = _calc(datos, parametros)
    ruta = escribir_formulario(PLANTILLA, tmp_path / "dep.xlsx", datos, liq)
    ws = _abrir(ruta)["Dependientes "]
    assert ws["C3"].value == 2
    assert ws["B7"].value == "Ana María"
    assert ws["B8"].value == "Luis"
    assert ws["B9"].value is None       # se limpian los nombres de ejemplo
    assert ws["B10"].value is None


def test_dependientes_solo_conteo_usa_marcadores(tmp_path, exogena_elizabeth, parametros):
    from src.motor_calculo import calcular as _calc
    datos = mapear_exogena_a_datos(exogena_elizabeth)
    datos.dependientes = 3              # sin nombres (flujo CLI antiguo)
    liq = _calc(datos, parametros)
    ruta = escribir_formulario(PLANTILLA, tmp_path / "dep2.xlsx", datos, liq)
    ws = _abrir(ruta)["Dependientes "]
    assert ws["B7"].value == "Dependiente 1"
    assert ws["B9"].value == "Dependiente 3"
    assert ws["B10"].value is None
