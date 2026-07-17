"""Las hojas de detalle deben traer los datos reales y no los ejemplos."""
import re
import warnings

import openpyxl
import pytest

from src.entrevista import mapear_exogena_a_datos
from src.excel_writer import escribir_formulario
from src.motor_calculo import calcular

from .conftest import PLANTILLA


@pytest.fixture(scope="module")
def wb_elizabeth(tmp_path_factory, exogena_elizabeth, parametros):
    datos = mapear_exogena_a_datos(exogena_elizabeth, parametros)
    datos.numero_anio_declaracion = 3
    datos.impuesto_neto_anio_anterior = 1_000_000
    liq = calcular(datos, parametros)
    ruta = escribir_formulario(PLANTILLA, tmp_path_factory.mktemp("out") / "s.xlsx",
                               datos, liq, exogena_elizabeth, parametros)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(ruta), datos, liq


def _constantes(ws, rango):
    vals = []
    for fila in ws[rango]:
        for c in fila:
            v = c.value
            if v is not None and not (isinstance(v, str) and v.startswith("=")):
                vals.append(v)
    return vals


def test_deudas_reales_sin_ejemplos(wb_elizabeth):
    wb, datos, _ = wb_elizabeth
    ws = wb["Deudas"]
    textos = [str(v) for v in _constantes(ws, "B2:C41")]
    assert not any("davivienda" in t.lower() for t in textos)   # ejemplo borrado
    assert any("BANCOLOMBIA" in t for t in textos)              # deuda real exógena
    valores = [v for v in _constantes(ws, "C2:C41") if isinstance(v, (int, float))]
    assert sum(valores) == pytest.approx(datos.deudas)


def test_pat_bruto_real_sin_ejemplos(wb_elizabeth):
    wb, datos, _ = wb_elizabeth
    ws = wb["Pat bruto"]
    textos = [str(v).lower() for v in _constantes(ws, "B5:D28")]
    assert not any("davivienda" in t for t in textos)
    assert any("cdt" in t for t in textos)                      # CDT BBVA real
    # los avalúos catastrales van como avalúo (col H) en activos fijos
    avaluos = [v for v in _constantes(ws, "H84:H100") if isinstance(v, (int, float))]
    assert sorted(avaluos) == [1_109_000, 3_934_000, 64_275_000]
    # ejemplos de vehículo/apto borrados
    textos_af = [str(v).lower() for v in _constantes(ws, "B84:B100")]
    assert not any("nissa" in t or "apto calle" in t for t in textos_af)


def test_retefuente_real(wb_elizabeth):
    wb, datos, _ = wb_elizabeth
    ws = wb["retefuente"]
    textos = [str(v) for v in _constantes(ws, "C3:C36")]
    assert not any("Itgs Sas" in t for t in textos)
    valores = [v for v in _constantes(ws, "D3:D36") if isinstance(v, (int, float))]
    assert sum(valores) == pytest.approx(datos.retenciones)


def test_anticipo_inputs(wb_elizabeth):
    wb, datos, _ = wb_elizabeth
    ws = wb["anticipo"]
    assert ws["E4"].value == datos.anticipo_anterior
    assert ws["E5"].value == 1_000_000
    assert ws["E6"].value == 3


def test_trabajo_ingresos_reales(wb_elizabeth):
    wb, datos, liq = wb_elizabeth
    ws = wb["R.trabajo y honorarios"]
    # G4 salarios + G6 cesantías consignadas + G12 resto = R32
    suma = sum(v for v in (ws["G4"].value, ws["G6"].value, ws["G12"].value)
               if isinstance(v, (int, float)))
    assert suma == pytest.approx(datos.trabajo.ingresos_brutos)
    # INCRNGO real (pensión + salud de la exógena)
    assert (ws["G21"].value or 0) + (ws["G23"].value or 0) == pytest.approx(datos.trabajo.incrngo)
    # ejemplo de 160M borrado
    assert ws["G4"].value != 160000000
    # renta exenta 25% escrita como valor del motor
    assert isinstance(ws["G62"].value, (int, float))
    assert ws["G62"].value == pytest.approx(liq.r(36) - round(datos.trabajo.otras_rentas_exentas))
    assert ws["H74"].value == 49799     # UVT actualizada


def test_capital_rendimientos_reales(wb_elizabeth):
    wb, datos, _ = wb_elizabeth
    ws = wb["R.capital"]
    rend = [v for v in _constantes(ws, "E3:E11") if isinstance(v, (int, float))]
    assert sum(rend) == pytest.approx(33_810_314)   # CDTs marcados R58|R59
    otros = [v for v in _constantes(ws, "E13:E19") if isinstance(v, (int, float))]
    assert sum(rend) + sum(otros) == pytest.approx(datos.capital.ingresos_brutos)
    # componente inflacionario con el % del config
    assert ws["F41"].value == "=F3*55.43%"
    # ejemplos borrados
    textos = [str(v).lower() for v in _constantes(ws, "C4:C11")]
    assert not any("caja social" in t for t in textos)


def test_no_laboral_real(wb_elizabeth):
    wb, datos, _ = wb_elizabeth
    ws = wb["R.no laboral y R gravables"]
    vals = [v for v in _constantes(ws, "F3:F16") if isinstance(v, (int, float))]
    assert sum(vals) == pytest.approx(datos.no_laboral.ingresos_brutos)
    textos = [str(v).lower() for v in _constantes(ws, "C3:C16")]
    assert not any("vehiculo" in t or "conyugal" in t for t in textos)


def test_dividendos_y_pensiones_limpios(wb_elizabeth):
    wb, _, _ = wb_elizabeth
    assert _constantes(wb["C divid."], "E3:E9") == []       # sin ejemplos (34M/25M)
    assert _constantes(wb["C divid."], "E17:E23") == []
    assert _constantes(wb["C.pensiones"], "E3:E8") == []


def test_gocas_limpia_y_uvt(wb_elizabeth):
    wb, _, _ = wb_elizabeth
    ws = wb["G.OCAS"]
    assert ws["D3"].value in (None, "")     # venta de ejemplo borrada
    assert ws["E89"].value == 49799


# ---------------------------------------------------------------- papeles pro

def test_anexo_exogena_trazabilidad(wb_elizabeth, exogena_elizabeth):
    """El Anexo lista TODAS las partidas sin recortes, con NIT, estado y total."""
    wb, _, _ = wb_elizabeth
    assert "Anexo Exógena" in wb.sheetnames
    ws = wb["Anexo Exógena"]
    # encabezados de la tabla
    cab = [ws.cell(row=4, column=j).value for j in range(1, 10)]
    assert cab[:4] == ["Fila", "Concepto reportado", "Informante", "NIT informante"]
    # tantas filas de datos como partidas (todas, incluidas las excluidas)
    n = len(exogena_elizabeth.partidas)
    filas = [ws.cell(row=5 + i, column=8).value for i in range(n)]
    assert all(f in ("Incluida", "Ajustada", "Excluida") for f in filas)
    # el total tomado = suma de partidas activas
    total = ws.cell(row=5 + n, column=6).value
    esperado = sum(p.valor for p in exogena_elizabeth.partidas_activas())
    assert total == pytest.approx(esperado)
    # interactivo: filtros y encabezado congelado
    assert ws.auto_filter.ref.startswith("A4:I")
    assert ws.freeze_panes == "A5"


def test_indice_navegable(wb_elizabeth):
    """El Índice existe, es la primera hoja y sus links apuntan a hojas reales."""
    wb, _, _ = wb_elizabeth
    assert wb.sheetnames[0] == "Índice"
    ws = wb["Índice"]
    links = [c.hyperlink.location for fila in ws["B5:B20"] for c in fila
             if c.hyperlink is not None]
    assert links, "el índice debe tener hipervínculos"
    for loc in links:
        hoja = loc.split("!")[0].strip("'")
        assert hoja in wb.sheetnames


def test_etiquetas_con_nit_del_informante(wb_elizabeth):
    """Las hojas de detalle muestran el NIT del tercero para poder cruzar."""
    wb, _, _ = wb_elizabeth
    textos = [str(v) for v in _constantes(wb["Deudas"], "B2:B41")]
    assert any(re.search(r"\(\d{6,}\)", t) for t in textos), \
        "las deudas deben traer el NIT del informante entre paréntesis"
